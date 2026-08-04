"""Jira 연동 라우터 (prefix `/jira`).

- 공통 설정(base_url 등)은 AppSetting key=`jira_integration` (관리자 전용 쓰기).
- 사용자별 PAT 는 `user_jira_credentials` 에 암호화 저장 (secret_box).
- 가져오기는 서버사이드에서 현재 사용자 PAT 로 실행 → work_items upsert (dedup by jira_issue_id).
"""
from __future__ import annotations

import asyncio
import html
import json
import logging
import re
from datetime import datetime
from html.parser import HTMLParser
from io import BytesIO
from pathlib import Path
from typing import Any, Optional

import openpyxl
import xlrd
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import PlainTextResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.app_setting import AppSetting
from app.models.user import User
from app.models.user_jira_credential import UserJiraCredential
from app.models.work_item import WorkItem
from app.auth.deps import get_current_user, require_admin, require_operator
from app.services import secret_box
from app.services import audit_logger
from app.services.jira_service import (
    ISSUE_FIELDS, JiraService, map_jira_issue, map_issue_type, parse_jira_dt, KANBAN_TO_CATEGORY,
    PEP_PRIORITY_TO_JIRA, strip_issue_key_prefix,
)
from app.services.confluence_service import ConfluenceService
from app.services.jira_sso_http import (
    CONFLUENCE_VERIFY_PATH, JIRA_VERIFY_PATH, diagnose_products, outbound_client_info,
    sso_login_products,
)
from app.services.jira_sso_service import capture_sso_session
from app.schemas.jira import (
    JiraConfig,
    JiraConfigUpdate,
    JiraCredentialStatus,
    JiraCredentialUpdate,
    JiraTestResult,
    JiraSsoLoginRequest,
    JiraSsoLoginResult,
    ConfluenceSearchItem,
    ConfluenceSearchResult,
    SsoDiagnoseEntry,
    SsoDiagnoseResult,
    JiraImportRequest,
    JiraImportResult,
    JiraImportItemPreview,
    JiraFieldChange,
    JiraExcelRow,
    JiraExcelImportResult,
    JiraExcelPasteRequest,
    JiraExcelSaveRequest,
    JiraPushRequest,
    JiraPushResult,
    JiraCreateRequest,
    JiraCreateResult,
    JiraDeleteResult,
    JiraUnlinkRequest,
    JiraUnlinkResult,
    JiraRelinkRequest,
    JiraRelinkResult,
    JiraMissingLink,
    JiraVerifyLinksResult,
    WeeklyReport,
    WeeklyReportRequest,
    WeeklyPublishRequest,
    WeeklyPublishResult,
    WeeklyReportSettings,
    ProvisionDefaults,
    ProvisionRequest,
    ProvisionResult,
)
from app.services import weekly_report_service
from app.services.user_settings import get_user_setting, set_user_setting
# 업무 삭제 권한 규칙(등록자/담당자/admin)을 게시판과 동일하게 적용하기 위해 재사용.
from app.routers.work_items import _assert_ownership as _assert_work_item_ownership

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/jira", tags=["jira"])

JIRA_SETTINGS_KEY = "jira_integration"
ASSIGNEES_KEY = "assignees"
DEFAULT_JIRA_SETTINGS = {
    "base_url": "",
    "enabled": False,
    "verify_tls": True,
    "default_project_key": None,
    # 같은 IdP 로 SSO 연동되는 Confluence Base URL (선택) — 설정 시 SSO 폼 로그인이
    # Jira 와 Confluence 세션을 한 번에 캡처한다.
    "confluence_base_url": "",
    # IdP 로그인 페이지 URL (선택) — 자동 탐색 실패 시 지정. 예:
    # https://login.example.com/sso/am/jira/login.jsp
    "sso_login_url": "",
    # IdP 로그인 폼의 계정 필드명 (선택) — 자동 추정 실패 시 지정 (예: empnum).
    "sso_username_field": "",
    # Jira Epic Link 커스텀 필드 ID (선택, 예: customfield_10008) — 주간보고 진척률의
    # Epic 축을 채운다. Server/DC 는 Epic Link 가 커스텀 필드라 인스턴스마다 ID 가 다르다.
    "jira_epic_field": "",
}


def _get_config(db: Session) -> dict:
    row = db.query(AppSetting).filter(AppSetting.key == JIRA_SETTINGS_KEY).first()
    value = dict(DEFAULT_JIRA_SETTINGS)
    if row and isinstance(row.value, dict):
        value.update(row.value)
    return value


def _build_assignee_resolver(db: Session):
    """Jira displayName → PEP 담당자 이름. 레지스트리(name) 와 대소문자 무시 매칭, 실패 시 원본."""
    try:
        row = db.query(AppSetting).filter(AppSetting.key == ASSIGNEES_KEY).first()
        registry = row.value if row and isinstance(row.value, list) else []
    except Exception:  # noqa: BLE001
        registry = []
    by_lower = {}
    for a in registry:
        if isinstance(a, dict) and a.get("name"):
            by_lower[str(a["name"]).strip().lower()] = str(a["name"]).strip()

    def _resolve(jira_name: str) -> str:
        return by_lower.get(jira_name.strip().lower(), jira_name.strip())

    return _resolve


def _build_assignee_roster(db: Session) -> dict[str, str]:
    """등록된 담당자 name → name (대소문자 무시 매칭용 조회 테이블)."""
    row = db.query(AppSetting).filter(AppSetting.key == ASSIGNEES_KEY).first()
    registry = row.value if row and isinstance(row.value, list) else []
    return {
        str(a["name"]).strip().lower(): str(a["name"]).strip()
        for a in registry if isinstance(a, dict) and a.get("name")
    }


def _match_excel_assignee(raw: str, roster_by_lower: dict[str, str]) -> tuple[Optional[str], bool]:
    """Jira Excel 의 담당자 셀("이름 회사")에서 이름을 추출해 담당자 레지스트리와 매칭.

    한글 이름은 공백 없는 한 토큰이 일반적이므로 첫 토큰을 이름 후보로 본다.
    실패하면 전체 문자열(회사명 없는 경우 대비)로 한 번 더 시도."""
    s = (raw or "").strip()
    if not s:
        return None, False
    first_token = s.split()[0]
    name = roster_by_lower.get(first_token.lower())
    if name:
        return name, True
    name = roster_by_lower.get(s.lower())
    if name:
        return name, True
    return first_token, False


_EXCEL_HEADER_ALIASES: dict[str, list[str]] = {
    "key": ["key", "issue key", "issue id"],
    "summary": ["summary"],
    "issue_type": ["issue type", "issuetype", "type"],
    "status": ["status"],
    "assignee": ["assignee"],
    "created": ["created"],
    "resolved": ["resolved"],
    "due_date": ["due date", "duedate"],
    "environment": ["environment"],
    "description": ["description"],
}


def _norm_excel_header(h) -> str:
    return re.sub(r"\s+", " ", str(h or "").strip().lower())


def _excel_cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M") if (v.hour or v.minute) else v.strftime("%Y-%m-%d")
    return str(v).strip()


_INLINE_TAG_RE = re.compile(r"<[^>]+>")


def _strip_inline_html(s: str) -> str:
    """Jira 의 HTML 기반 내보내기는 Description/Environment 같은 rich-text 필드 안에
    이스케이프된 HTML(예: `&lt;p dir="auto"&gt;...&lt;/p&gt;`)이 들어있는 경우가 있다 —
    html.parser 가 엔티티를 이미 복원해버려 `<p dir="auto">...` 처럼 태그가 그대로 텍스트로
    노출된다. 태그를 제거하고 남은 텍스트만 정리해 돌려준다."""
    if not s:
        return s
    text = html.unescape(s)
    text = _INLINE_TAG_RE.sub(" ", text)
    return re.sub(r"\s+", " ", text).strip()


# Jira 날짜 표시 포맷 후보 — HTML 기반 내보내기는 날짜를 문자열로 담고(예: "11/Jun/26 10:31 AM"),
# xlsx/xls 는 네이티브 datetime 셀일 수 있어(이미 _excel_cell_str 이 "YYYY-MM-DD[ HH:MM]" 로
# 정규화) 두 경우 모두 커버한다.
_EXCEL_DATE_FORMATS = (
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d",
    "%d/%b/%y %I:%M %p",
    "%d/%b/%Y %I:%M %p",
    "%d/%b/%y",
    "%d/%b/%Y",
    "%m/%d/%Y %I:%M %p",
    "%m/%d/%y %I:%M %p",
    "%m/%d/%Y",
)


def _parse_excel_date(v) -> Optional[datetime]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in _EXCEL_DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _excel_date_only(v) -> str:
    """날짜 셀을 시간 없이 "YYYY-MM-DD" 로만 표시(Created 등은 시간까지는 불필요하다는
    요청). 알려진 포맷으로 파싱되면 그 날짜를, 실패하면 첫 공백 앞부분만 보수적으로 남긴다."""
    dt = _parse_excel_date(v)
    if dt:
        return dt.strftime("%Y-%m-%d")
    s = str(v or "").strip()
    return s.split(" ")[0] if s else ""


_DONE_STATUS_WORDS = ("done", "closed", "resolved", "complete", "완료", "해결", "종료", "닫힘")
_INPROGRESS_STATUS_WORDS = ("progress", "review", "진행", "검토", "처리중", "처리 중")


def _map_excel_status_to_kanban(status_name: str) -> str:
    """엑셀에는 Jira 의 statusCategory(new/indeterminate/done) 가 없고 상태명 문자열만
    있으므로, 흔한 표기를 텍스트 매칭으로 추정한다(라이브 JQL 가져오기의
    `map_status_category` 보다 느슨함 — 커스텀 워크플로 상태명은 기본값 todo 로 떨어진다)."""
    s = (status_name or "").strip().lower()
    if any(w in s for w in _DONE_STATUS_WORDS):
        return "done"
    if any(w in s for w in _INPROGRESS_STATUS_WORDS):
        return "in_progress"
    return "todo"


def _read_xls_rows(raw: bytes):
    """.xls(레거시 바이너리) 워크북의 첫 시트를 openpyxl 의 iter_rows(values_only=True) 와
    동일한 형태(행마다 값 튜플, 날짜 셀은 datetime)로 반환. 워크북을 여는 단계(가장 흔한 실패
    지점)는 이 함수 호출 시점에 즉시 실행돼 호출부의 try/except 로 잡힌다 — 아래 _rows() 의
    행 순회만 지연 평가된다."""
    wb = xlrd.open_workbook(file_contents=raw)
    sheet = wb.sheet_by_index(0)

    def _rows():
        for row_idx in range(sheet.nrows):
            values = []
            for c in sheet.row(row_idx):
                if c.ctype == xlrd.XL_CELL_DATE:
                    try:
                        values.append(xlrd.xldate_as_datetime(c.value, wb.datemode))
                    except Exception:  # noqa: BLE001
                        values.append(c.value)
                elif c.ctype == xlrd.XL_CELL_EMPTY:
                    values.append(None)
                else:
                    values.append(c.value)
            yield tuple(values)

    return _rows()


def _looks_like_html(raw: bytes) -> bool:
    """Jira '엑셀(전체 필드)' 내보내기는 확장자가 `.xls` 지만 실제 내용은 HTML 테이블이다
    (구버전 Excel 이 확장자만 보고 열어주던 호환 방식). xlrd 는 진짜 OLE2 바이너리만 지원해
    이런 파일에 'Expected BOF record' 로 실패하므로, 업로드 시점에 먼저 감지한다."""
    head = raw[:1024].lstrip().lower()
    return head.startswith((b"<html", b"<!doctype", b"<?xml")) or b"<table" in raw[:4096].lower()


class _JiraHtmlTableExtractor(HTMLParser):
    """HTML 문서 안의 모든 <table>(중첩 포함) 을 각각 독립된 행 목록으로 추출.
    표준 라이브러리만 사용하는 최소 파서.

    과거에는 "첫 번째 <table> 만" 사용했는데, Jira 내보내기가 요약/메타 정보를 담은 작은
    표를 실제 이슈 목록 표보다 앞에 두거나(형제 테이블), 레이아웃용 바깥 테이블 <td> 안에
    실제 이슈 표를 중첩시키는 경우 모두 있어 — 첫 테이블만 보면 진짜 데이터 표를 통째로
    놓치고 "필수 컬럼을 찾을 수 없음" 오류가 났다. 테이블 스택으로 중첩을 추적해 발견되는
    모든 <table> 을 순서대로 `self.tables` 에 쌓고, 호출부가 각 표를 순회하며 Key/Summary
    헤더를 가진 표를 찾는다(아래 `_find_header` 참고)."""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.tables: list[list[list[str]]] = []
        self._table_stack: list[list[list[str]]] = []  # 진행 중인 표(중첩 가능) 스택
        self._row_stack: list[list[str]] = []           # 스택 위치별 현재 행 버퍼
        self._in_cell = False
        self._cell_buf: list[str] = []

    def handle_starttag(self, tag: str, attrs) -> None:  # noqa: ARG002
        if tag == "table":
            self._table_stack.append([])
            self._row_stack.append([])
        elif tag == "tr" and self._table_stack:
            self._row_stack[-1] = []
        elif tag in ("td", "th") and self._table_stack:
            self._in_cell = True
            self._cell_buf = []
        elif self._in_cell and tag == "br":
            self._cell_buf.append("\n")

    def handle_endtag(self, tag: str) -> None:
        if tag in ("td", "th") and self._in_cell:
            self._in_cell = False
            self._row_stack[-1].append("".join(self._cell_buf).strip())
        elif tag == "tr" and self._table_stack:
            self._table_stack[-1].append(self._row_stack[-1])
        elif tag == "table" and self._table_stack:
            rows = self._table_stack.pop()
            self._row_stack.pop()
            self.tables.append(rows)

    def handle_data(self, data: str) -> None:
        if self._in_cell:
            self._cell_buf.append(data)


def _read_html_tables(raw: bytes) -> list[list[tuple]]:
    """Jira 의 HTML 기반 '가짜 .xls' 내보내기를 파싱해 문서 안의 모든 표를 반환한다
    (표 하나 = 행 튜플 리스트, 완전히 빈 행은 제거). 인코딩은 UTF-8 우선, 실패 시
    한글 환경에서 흔한 CP949/EUC-KR 로 재시도."""
    text: Optional[str] = None
    for enc in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            text = raw.decode(enc)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    if text is None:
        text = raw.decode("utf-8", errors="replace")
    parser = _JiraHtmlTableExtractor()
    parser.feed(text)
    return [
        [tuple(r) for r in table if any(v for v in r)]
        for table in parser.tables
    ]


def _normalize_cookie_header(raw: str) -> str:
    """수동 등록 세션 쿠키 입력 정규화.

    사용자가 자주 틀리는 두 가지를 흡수한다:
     - DevTools 에서 헤더째 복사해 앞에 `Cookie:` 가 붙은 경우 → 제거
     - **값만** 붙여넣은 경우(예: `A1B2C3...`) → `Cookie: A1B2C3` 는 이름이 없어 서버가
       무시하므로 익명 취급되어 401 이 된다. `name=value` 쌍이 하나도 없으면 Jira/Confluence
       의 세션 쿠키 이름인 `JSESSIONID=` 를 붙여준다."""
    s = (raw or "").strip()
    if s.lower().startswith("cookie:"):
        s = s.split(":", 1)[1].strip()
    if "=" not in s:
        return f"JSESSIONID={s}" if s else s
    return s


def _user_credential(db: Session, username: str) -> tuple[Optional[str], str]:
    """사용자별 Jira 자격 (복호화된 secret, auth_type) 반환. 미등록/복호화 실패 시 (None, 'pat')."""
    cred = db.query(UserJiraCredential).filter(UserJiraCredential.username == username).first()
    if not cred:
        return None, "pat"
    auth_type = (getattr(cred, "auth_type", None) or "pat").strip().lower()
    try:
        return secret_box.decrypt(cred.token_encrypted), auth_type
    except ValueError:
        return None, auth_type


# ── 공통 설정 ──────────────────────────────────────────────────────────────────
@router.get("/config", response_model=JiraConfig)
def get_config(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return JiraConfig(**_get_config(db))


@router.put("/config", response_model=JiraConfig)
def update_config(
    payload: JiraConfigUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    row = db.query(AppSetting).filter(AppSetting.key == JIRA_SETTINGS_KEY).first()
    current = dict(DEFAULT_JIRA_SETTINGS)
    if row and isinstance(row.value, dict):
        current.update(row.value)
    data = payload.model_dump(exclude_unset=True)
    if "base_url" in data and data["base_url"] is not None:
        current["base_url"] = data["base_url"].rstrip("/")
    if "confluence_base_url" in data and data["confluence_base_url"] is not None:
        current["confluence_base_url"] = data["confluence_base_url"].strip().rstrip("/")
    if "sso_login_url" in data and data["sso_login_url"] is not None:
        # IdP 로그인 URL 은 쿼리스트링(goto 등)을 포함할 수 있어 rstrip('/') 하지 않는다.
        current["sso_login_url"] = data["sso_login_url"].strip()
    if "sso_username_field" in data and data["sso_username_field"] is not None:
        current["sso_username_field"] = data["sso_username_field"].strip()
    if "jira_epic_field" in data and data["jira_epic_field"] is not None:
        current["jira_epic_field"] = data["jira_epic_field"].strip()
    for k in ("enabled", "verify_tls", "default_project_key"):
        if k in data and data[k] is not None:
            current[k] = data[k]
    if row:
        row.value = current
    else:
        db.add(AppSetting(key=JIRA_SETTINGS_KEY, value=current))
    db.commit()
    return JiraConfig(**current)


# ── 사용자별 자격증명 ──────────────────────────────────────────────────────────
@router.get("/credential", response_model=JiraCredentialStatus)
def get_credential(db: Session = Depends(get_db), actor: User = Depends(get_current_user)):
    cred = db.query(UserJiraCredential).filter(UserJiraCredential.username == actor.username).first()
    if not cred:
        return JiraCredentialStatus(configured=False)
    return JiraCredentialStatus(
        configured=True,
        auth_type=(getattr(cred, "auth_type", None) or "pat"),
        jira_account=cred.jira_account,
        last_verified_at=cred.last_verified_at,
        has_sso_login=bool(getattr(cred, "sso_login_encrypted", None)),
        has_confluence=bool(getattr(cred, "confluence_cookie_encrypted", None)),
    )


@router.put("/credential", response_model=JiraCredentialStatus)
def save_credential(
    payload: JiraCredentialUpdate,
    db: Session = Depends(get_db),
    actor: User = Depends(get_current_user),
):
    token = (payload.token or "").strip()
    if not token:
        raise HTTPException(status_code=422, detail="인증 값을 입력하세요 (PAT 또는 세션 쿠키).")
    auth_type = (payload.auth_type or "pat").strip().lower()
    # 'sso' 는 로컬 도우미(jira_sso_helper.py)가 본인 PC 에서 캡처한 세션 쿠키를 등록하는 경로 —
    # REST 처리(JiraService)는 cookie 와 동일하고, UI 배지만 SSO 로 구분 표시된다.
    if auth_type not in ("pat", "cookie", "sso"):
        raise HTTPException(status_code=422, detail="auth_type 은 'pat'/'cookie'/'sso' 여야 합니다.")
    if auth_type in ("cookie", "sso"):
        token = _normalize_cookie_header(token)
    enc = secret_box.encrypt(token)
    cred = db.query(UserJiraCredential).filter(UserJiraCredential.username == actor.username).first()
    if cred:
        cred.token_encrypted = enc
        cred.auth_type = auth_type
        # 인증 값을 새로 저장하면 이전 검증 시각은 무의미 — 초기화해 재검증을 유도.
        cred.last_verified_at = None
        if payload.jira_account is not None:
            cred.jira_account = payload.jira_account
    else:
        cred = UserJiraCredential(
            username=actor.username, token_encrypted=enc, auth_type=auth_type,
            jira_account=payload.jira_account,
        )
        db.add(cred)
    db.commit()
    db.refresh(cred)
    return JiraCredentialStatus(
        configured=True, auth_type=cred.auth_type,
        jira_account=cred.jira_account, last_verified_at=cred.last_verified_at,
    )


@router.delete("/credential", status_code=status.HTTP_204_NO_CONTENT)
def delete_credential(db: Session = Depends(get_db), actor: User = Depends(get_current_user)):
    cred = db.query(UserJiraCredential).filter(UserJiraCredential.username == actor.username).first()
    if cred:
        db.delete(cred)
        db.commit()
    return None


@router.post("/test", response_model=JiraTestResult)
async def test_connection(db: Session = Depends(get_db), actor: User = Depends(get_current_user)):
    cfg = _get_config(db)
    if not cfg.get("base_url"):
        return JiraTestResult(ok=False, detail="관리자가 Jira URL 을 설정하지 않았습니다.")
    # 세션 만료(401) 시 저장된 SSO 로그인 정보로 자동 재로그인 포함.
    svc, res = await _jira_service_verified(db, actor, cfg)
    if svc is None:
        return JiraTestResult(ok=False, detail="내 Jira 인증(PAT 또는 세션 쿠키)이 등록되지 않았습니다.")
    if res.get("status") == "ok":
        cred = db.query(UserJiraCredential).filter(UserJiraCredential.username == actor.username).first()
        if cred:
            cred.last_verified_at = datetime.utcnow()
            if res.get("account") and not cred.jira_account:
                cred.jira_account = res.get("display_name") or res.get("account")
            db.commit()
        return JiraTestResult(ok=True, detail="연결 정상", display_name=res.get("display_name"))
    return JiraTestResult(ok=False, detail=res.get("detail", "연결 실패"))


# ── SSO 자동 로그인 ────────────────────────────────────────────────────────────
# 두 가지 실행 모드:
#  - ID/PW 폼 로그인 (기본, K8s 배포용) — 파드 안에서 httpx 로 SSO 리다이렉트 체인을 따라가
#    폼을 제출하고 쿠키를 캡처한다. 브라우저 불필요 (jira_sso_http.sso_login_products).
#    관리자가 Confluence URL 을 설정했으면 같은 IdP 세션으로 Confluence 세션까지 한 번에
#    캡처한다. save_login 옵트인 시 로그인 정보를 암호화 저장해 원클릭/자동 재로그인 지원.
#  - Playwright 헤디드 로그인 (레거시) — 백엔드 호스트에 화면이 있는 소스 실행 배포 전용.


def _sso_products(cfg: dict) -> list[dict]:
    """SSO 폼 로그인 대상 제품 목록 — 첫 항목(Jira)이 주 제품, Confluence 는 설정 시에만.

    관리자가 IdP 로그인 URL 을 지정했으면 Jira 진입점 맨 앞에 놓는다(자동 탐색 실패 대비).
    Confluence 는 Jira 로그인으로 이미 IdP 세션이 생긴 뒤라 제품 진입만으로 통과한다."""
    products = [{
        "key": "jira", "label": "Jira",
        "base_url": cfg.get("base_url", ""), "verify_path": JIRA_VERIFY_PATH,
        "sso_login_url": (cfg.get("sso_login_url") or "").strip(),
        "username_field": (cfg.get("sso_username_field") or "").strip(),
    }]
    conf_url = (cfg.get("confluence_base_url") or "").strip()
    if conf_url:
        products.append({
            "key": "confluence", "label": "Confluence",
            "base_url": conf_url, "verify_path": CONFLUENCE_VERIFY_PATH,
            "username_field": (cfg.get("sso_username_field") or "").strip(),
        })
    return products


async def _sso_relogin(db: Session, actor: User, cfg: dict) -> Optional[str]:
    """저장된 SSO 로그인 정보(옵트인)로 파드 내 폼 재로그인 → 새 세션 쿠키 저장(커밋 포함).

    Jira 세션이 만료된 API 호출 경로에서 자동으로 불린다. Confluence 가 설정돼 있으면
    그 세션도 함께 갱신된다. 성공 시 새 Jira Cookie 헤더, 실패/저장정보 없음이면 None."""
    cred = db.query(UserJiraCredential).filter(UserJiraCredential.username == actor.username).first()
    if not (cred and getattr(cred, "sso_login_encrypted", None)):
        return None
    try:
        saved = json.loads(secret_box.decrypt(cred.sso_login_encrypted))
    except Exception:  # noqa: BLE001 - SECRET_KEY 교체 등으로 복호 실패
        return None
    sso_username, sso_password = saved.get("username"), saved.get("password")
    if not (sso_username and sso_password):
        return None
    result = await sso_login_products(
        _sso_products(cfg), sso_username, sso_password,
        verify_tls=bool(cfg.get("verify_tls", True)),
    )
    if result.get("status") != "ok":
        logger.info("Jira SSO auto re-login failed for %s: %s",
                    actor.username, result.get("detail"))
        return None
    cred.token_encrypted = secret_box.encrypt(result["cookie_header"])
    cred.auth_type = "sso"
    cred.last_verified_at = datetime.utcnow()
    if result.get("display_name") and not cred.jira_account:
        cred.jira_account = result["display_name"]
    conf = (result.get("products") or {}).get("confluence")
    if conf and conf.get("status") == "ok":
        cred.confluence_cookie_encrypted = secret_box.encrypt(conf["cookie_header"])
    db.commit()
    audit_logger.record(
        db, action="work_item.jira_sso_relogin", actor=actor,
        target_type="user_jira_credential", target_id=None,
        details={"auto": True, "confluence": bool(conf and conf.get("status") == "ok")},
    )
    logger.info("Jira SSO auto re-login ok for %s", actor.username)
    return result["cookie_header"]


async def _jira_service_verified(db: Session, actor: User, cfg: dict) -> tuple[Optional[JiraService], dict]:
    """사용자 자격으로 JiraService 생성 + myself 1회 검증.

    세션 만료(401)이고 저장된 SSO 로그인 정보가 있으면 자동 재로그인 후 새 세션으로 재시도 —
    사용자는 최초 1회만 로그인하면 세션이 끊겨도 무중단으로 이어진다.
    반환 (svc, myself 결과). 인증 미등록이면 (None, error dict)."""
    token, auth_type = _user_credential(db, actor.username)
    if not token:
        return None, {"status": "error",
                      "detail": "내 Jira 인증(PAT 또는 세션 쿠키)이 등록되지 않았습니다."}
    verify = bool(cfg.get("verify_tls", True))
    svc = JiraService(cfg["base_url"], token, auth_type=auth_type, verify=verify)
    res = await svc.myself()
    if res.get("auth_failed") and auth_type in ("cookie", "sso"):
        new_cookie = await _sso_relogin(db, actor, cfg)
        if new_cookie:
            svc = JiraService(cfg["base_url"], new_cookie, auth_type="sso", verify=verify)
            res = await svc.myself()
    return svc, res


def _user_confluence_cookie(db: Session, username: str) -> Optional[str]:
    cred = db.query(UserJiraCredential).filter(UserJiraCredential.username == username).first()
    if not (cred and getattr(cred, "confluence_cookie_encrypted", None)):
        return None
    try:
        return secret_box.decrypt(cred.confluence_cookie_encrypted)
    except ValueError:
        return None


async def _confluence_service_verified(
    db: Session, actor: User, cfg: dict
) -> tuple[Optional[ConfluenceService], dict]:
    """Confluence 세션으로 서비스 생성 + current_user 검증.

    세션이 없거나 만료됐으면 저장된 SSO 로그인 정보로 자동 (재)로그인해 확보한다
    (Jira 세션도 함께 갱신됨). 반환 (svc, current_user 결과) — 확보 실패 시 (None, error)."""
    conf_url = (cfg.get("confluence_base_url") or "").strip()
    if not conf_url:
        return None, {"status": "error", "detail": "관리자가 Confluence URL 을 설정하지 않았습니다."}
    verify = bool(cfg.get("verify_tls", True))
    cookie = _user_confluence_cookie(db, actor.username)
    if cookie:
        svc = ConfluenceService(conf_url, cookie, auth_type="sso", verify=verify)
        res = await svc.current_user()
        if not res.get("auth_failed"):
            return svc, res
    # Confluence 전용 세션이 없으면 **Jira 자격으로 폴백**한다 — SiteMinder 류 SSO 는
    # SMSESSION 을 상위 도메인에 발급하므로 같은 쿠키로 Confluence 도 통하는 경우가 많다.
    # (수동으로 세션 쿠키만 등록한 사용자는 Confluence 쿠키가 아예 없다.)
    jira_token, jira_auth = _user_credential(db, actor.username)
    if jira_token and jira_auth in ("cookie", "sso"):
        svc = ConfluenceService(conf_url, jira_token, auth_type="sso", verify=verify)
        res = await svc.current_user()
        if res.get("status") == "ok":
            # 통했으면 Confluence 세션으로 승격 저장 — 다음부터 바로 쓰인다.
            cred = db.query(UserJiraCredential).filter(
                UserJiraCredential.username == actor.username).first()
            if cred:
                cred.confluence_cookie_encrypted = secret_box.encrypt(jira_token)
                db.commit()
            return svc, res
    if await _sso_relogin(db, actor, cfg):
        cookie = _user_confluence_cookie(db, actor.username)
        if cookie:
            svc = ConfluenceService(conf_url, cookie, auth_type="sso", verify=verify)
            return svc, await svc.current_user()
    return None, {"status": "error",
                  "detail": "Confluence 세션이 없습니다 — 'SSO 자동 로그인'으로 세션을 캡처하세요."}


@router.post("/sso/login", response_model=JiraSsoLoginResult)
async def sso_login(
    payload: Optional[JiraSsoLoginRequest] = None,
    db: Session = Depends(get_db),
    actor: User = Depends(get_current_user),
):
    """SSO 로그인을 서버가 대신 수행하고 세션 쿠키를 자동 저장한다 (docstring 위 주석 참고)."""
    cfg = _get_config(db)
    base_url = cfg.get("base_url", "")
    if not base_url:
        return JiraSsoLoginResult(ok=False, detail="관리자가 Jira URL 을 설정하지 않았습니다.")
    verify_tls = bool(cfg.get("verify_tls", True))

    cred = db.query(UserJiraCredential).filter(UserJiraCredential.username == actor.username).first()
    sso_username: Optional[str] = None
    sso_password: Optional[str] = None

    if payload and payload.use_saved:
        if not (cred and cred.sso_login_encrypted):
            return JiraSsoLoginResult(ok=False, detail="저장된 SSO 로그인 정보가 없습니다. 아이디/비밀번호를 입력하세요.")
        try:
            saved = json.loads(secret_box.decrypt(cred.sso_login_encrypted))
            sso_username, sso_password = saved.get("username"), saved.get("password")
        except Exception:  # noqa: BLE001 - SECRET_KEY 교체 등으로 복호 실패
            return JiraSsoLoginResult(ok=False, detail="저장된 로그인 정보를 복호화할 수 없습니다. 다시 입력해 저장하세요.")
    elif payload and (payload.username or payload.password):
        if not (payload.username and payload.password):
            return JiraSsoLoginResult(ok=False, detail="SSO 아이디와 비밀번호를 모두 입력하세요.")
        sso_username, sso_password = payload.username.strip(), payload.password

    if sso_username and sso_password:
        # Jira(주 제품) + Confluence(설정 시) 를 한 IdP 세션으로 연속 로그인.
        result = await sso_login_products(
            _sso_products(cfg), sso_username, sso_password, verify_tls=verify_tls,
        )
    else:
        # 레거시 — 블로킹 Playwright 헤디드 로그인(사용자가 서버 브라우저에서 완료할 때까지 대기).
        result = await asyncio.to_thread(capture_sso_session, base_url, verify_tls=verify_tls)
    if result.get("status") != "ok":
        return JiraSsoLoginResult(ok=False, detail=result.get("detail", "SSO 로그인 실패"))

    cookie_header = result["cookie_header"]
    # 캡처한 쿠키로 실제 REST 접근이 되는지 검증(사용자 권한 확인).
    svc = JiraService(base_url, cookie_header, auth_type="sso", verify=verify_tls)
    verified = await svc.myself()
    if verified.get("status") != "ok":
        return JiraSsoLoginResult(
            ok=False,
            detail=f"로그인은 감지됐으나 백엔드에서 세션 검증에 실패했습니다: {verified.get('detail', '')} "
                   "(자체서명 인증서면 공통설정 'TLS 인증서 검증' 해제, 백엔드→Jira 네트워크 확인).",
        )

    display = verified.get("display_name") or result.get("display_name")
    account = result.get("account") or verified.get("account")
    enc = secret_box.encrypt(cookie_header)
    now = datetime.utcnow()
    if cred:
        cred.token_encrypted = enc
        cred.auth_type = "sso"
        cred.jira_account = display or account or cred.jira_account
        cred.last_verified_at = now
    else:
        cred = UserJiraCredential(
            username=actor.username, token_encrypted=enc, auth_type="sso",
            jira_account=display or account, last_verified_at=now,
        )
        db.add(cred)
    # 옵트인 — 로그인 정보 저장(원클릭/자동 재로그인용). 저장 없이 성공한 로그인은 기존 값을 유지.
    if payload and payload.save_login and sso_username and sso_password:
        cred.sso_login_encrypted = secret_box.encrypt(
            json.dumps({"username": sso_username, "password": sso_password})
        )
    # Confluence 동시 로그인 결과 — 설정된 경우에만 존재. 실패해도 Jira 로그인은 유효.
    conf = (result.get("products") or {}).get("confluence")
    confluence_ok: Optional[bool] = None
    confluence_detail = ""
    if conf is not None:
        confluence_ok = conf.get("status") == "ok"
        if confluence_ok:
            cred.confluence_cookie_encrypted = secret_box.encrypt(conf["cookie_header"])
        else:
            confluence_detail = conf.get("detail", "Confluence 로그인 실패")
    db.commit()
    audit_logger.record(
        db, action="work_item.jira_sso_login", actor=actor,
        target_type="user_jira_credential", target_id=None,
        details={"account": account, "method": "form" if sso_username else "browser",
                 "confluence": confluence_ok},
    )
    return JiraSsoLoginResult(
        ok=True,
        detail="SSO 로그인 완료 — 세션이 저장되었습니다."
               + (" (Confluence 포함)" if confluence_ok else ""),
        jira_account=account, display_name=display,
        confluence_ok=confluence_ok, confluence_detail=confluence_detail,
    )


@router.post("/sso/diagnose", response_model=SsoDiagnoseResult)
async def sso_diagnose(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    """SSO 진단 — 자격 없이 각 진입 경로를 GET 해 **파드가 보는 로그인 페이지**를 보고한다.

    폐쇄망 IdP 는 밖에서 열어볼 수 없어 로그인 실패 원인을 추측하기 어렵다. 최종 URL(IdP 로
    넘어갔는지), 폼/password 입력 개수(폼이 정말 없는지), client_redirect(JS·meta 중계인지),
    www_authenticate(Negotiate/Basic 인지)를 보면 어디서 끊겼는지 바로 판별된다."""
    cfg = _get_config(db)
    if not cfg.get("base_url"):
        return SsoDiagnoseResult(ok=False, detail="관리자가 Jira URL 을 설정하지 않았습니다.")
    rows = await diagnose_products(_sso_products(cfg), verify_tls=bool(cfg.get("verify_tls", True)))
    entries = [SsoDiagnoseEntry(**r) for r in rows]
    found = next((e for e in entries if e.password_inputs > 0), None)
    hints = sorted({h for e in entries for h in e.crypto_hints})
    if found and hints:
        # 클라이언트에서 자격을 가공하는 페이지는 평문 POST 로 절대 인증되지 않는다.
        detail = (
            f"로그인 폼은 찾았지만({found.final_url}) 이 페이지는 **브라우저에서 자격을 가공**하는 "
            f"것으로 보입니다: {', '.join(hints)}. 이 경우 서버측 폼 로그인은 원리상 실패하므로 "
            "'로컬 도우미'(내 PC 브라우저) 방식이나 PAT/세션 쿠키 등록을 사용하세요."
        )
    elif found:
        detail = f"로그인 폼 발견 — {found.final_url} (password 입력 {found.password_inputs}개). SSO 로그인을 시도해도 됩니다."
    else:
        detail = ("어느 진입 경로에서도 password 입력을 찾지 못했습니다. 아래 표의 final_url 이 "
                  "IdP 주소가 아니면 리다이렉트가 안 걸린 것이고, IdP 인데도 폼이 0 이면 JS 렌더링입니다. "
                  "브라우저에서 확인한 IdP 로그인 페이지 주소를 공통 설정의 'IdP 로그인 URL' 에 넣어보세요.")
    who = outbound_client_info(cfg.get("base_url", ""))
    return SsoDiagnoseResult(
        ok=bool(found) and not hints, detail=detail, entries=entries,
        pod_hostname=who.get("hostname", ""), pod_source_ip=who.get("source_ip", ""),
    )


# 백엔드가 K8s/컨테이너 배포라 파드에서 브라우저를 못 띄우는 환경용 — 사용자가 본인 PC 에서
# 실행해 SSO 세션을 캡처·등록하는 도우미 스크립트(이미지에 동봉)를 내려준다.
_SSO_HELPER_PATH = Path(__file__).resolve().parent.parent / "resources" / "jira_sso_helper.py"


@router.get("/sso/helper")
def download_sso_helper(_: User = Depends(get_current_user)):
    try:
        content = _SSO_HELPER_PATH.read_text(encoding="utf-8")
    except OSError:
        raise HTTPException(status_code=404, detail="도우미 스크립트가 이 배포 이미지에 포함되지 않았습니다.")
    return PlainTextResponse(
        content,
        media_type="text/x-python",
        headers={"Content-Disposition": 'attachment; filename="jira_sso_helper.py"'},
    )


def _jql_quote(v: str) -> str:
    """JQL 문자열 리터럴 — 역슬래시/따옴표 이스케이프."""
    return (v or "").replace("\\", "\\\\").replace('"', '\\"')


def _build_filter_jql(payload: JiraImportRequest) -> tuple[str, str]:
    """프로젝트/라벨/컴포넌트/상태/담당자/변경일 조건을 AND 로 묶어 JQL 을 조립한다.

    빈 항목은 무시하며, 여러 값은 `IN (...)` 으로 OR 처리한다.
    반환 (jql, error) — 조건이 하나도 없으면 error 를 채운다."""
    clauses: list[str] = []
    # 프로젝트도 쉼표로 여러 개 지정 가능 — 컴포넌트/라벨과 동일하게 개별·조합 모두 지원.
    projects = [x.strip() for x in (payload.project_key or "").split(",") if x.strip()]
    if len(projects) == 1:
        clauses.append(f'project = "{_jql_quote(projects[0])}"')
    elif projects:
        joined = ", ".join(f'"{_jql_quote(x)}"' for x in projects)
        clauses.append(f"project IN ({joined})")
    labels = [x.strip() for x in payload.labels if x and x.strip()]
    if labels:
        joined = ", ".join(f'"{_jql_quote(x)}"' for x in labels)
        clauses.append(f"labels IN ({joined})")
    comps = [x.strip() for x in payload.components if x and x.strip()]
    if comps:
        joined = ", ".join(f'"{_jql_quote(x)}"' for x in comps)
        clauses.append(f"component IN ({joined})")
    statuses = [x.strip() for x in payload.statuses if x and x.strip()]
    if statuses:
        joined = ", ".join(f'"{_jql_quote(x)}"' for x in statuses)
        clauses.append(f"status IN ({joined})")
    assignee = (payload.assignee or "").strip()
    if assignee:
        clauses.append(
            "assignee = currentUser()" if assignee.lower() == "currentuser()"
            else f'assignee = "{_jql_quote(assignee)}"'
        )
    days = payload.updated_since_days
    if days and days > 0:
        clauses.append(f"updated >= -{int(days)}d")
    if not clauses:
        return "", "조건을 하나 이상 지정하세요 (프로젝트/라벨/컴포넌트/상태/담당자)."
    return " AND ".join(clauses) + " ORDER BY updated DESC", ""


# 재가져오기 시 비교할 Jira 소유 필드 — (WorkItem 속성, 표시 라벨).
_SYNC_FIELDS: tuple[tuple[str, str], ...] = (
    ("title", "제목"),
    ("content", "내용"),
    ("kanban_status", "진행 상태"),
    ("priority", "우선순위"),
    ("jira_status", "Jira 상태"),
    ("category", "업무 분류"),
    ("jira_issue_type", "이슈 종류"),
    ("jira_epic", "Epic"),
    ("jira_parent_key", "상위 이슈"),
    ("jira_components", "컴포넌트"),
    ("jira_labels", "라벨"),
    ("confluence_url", "Confluence 링크"),
)


# Jira 가 소유하는 필드 — 가져올 때마다 무조건 덮어쓴다.
_JIRA_OWNED_ATTRS: tuple[str, ...] = (
    "title", "content", "kanban_status", "priority",
    "jira_issue_key", "jira_url", "jira_status", "jira_status_category",
    "jira_updated_at", "jira_issue_type", "jira_parent_key", "jira_parent_summary",
    "jira_components", "jira_labels",
)


def _jira_sync_values(existing: Optional[WorkItem], fields: dict) -> dict:
    """Jira 최신값 → **실제로 업무에 쓸 값** 만 추린 dict.

    보존 규칙(비었으면 기존 값 유지 / 로컬 편집 존중)을 이 함수 한 곳에만 두고
    변경 diff(`_diff_existing`)와 적용(`_apply_jira_fields`)이 똑같은 결과를 보게 한다.
    두 쪽이 어긋나면 "덮어쓰지 않는 필드"가 매번 변경으로 잡혀 `unchanged` 판정이
    영원히 나오지 않는다(재가져오기 때마다 update 로 집계되는 버그)."""
    out: dict[str, Any] = {a: fields[a] for a in _JIRA_OWNED_ATTRS if a in fields}
    # Epic 은 값이 있을 때만 — `jira_epic_field` 미설정 배포에서 기존 Epic 이 날아가지 않게.
    for attr in ("jira_epic", "jira_epic_key", "jira_epic_summary"):
        if fields.get(attr):
            out[attr] = fields[attr]
    # 업무 분류는 Jira component 를 찾았을 때만 갱신 — component 없는 이슈가 사용자가
    # 정해둔 분류를 폴백값("Jira")으로 되돌리지 않도록 한다.
    if fields.get("jira_components") and fields.get("category"):
        out["category"] = fields["category"]
    # Confluence 링크는 **비어 있을 때만** 채운다 — 사용자가 직접 넣은 링크를 덮지 않는다.
    if fields.get("confluence_url") and not (getattr(existing, "confluence_url", "") or "").strip():
        out["confluence_url"] = fields["confluence_url"]
    return out


def _diff_existing(existing: WorkItem, fields: dict) -> list[JiraFieldChange]:
    """기존 업무와 Jira 최신값의 차이 — 확인 팝업에 그대로 보여준다."""
    values = _jira_sync_values(existing, fields)
    out: list[JiraFieldChange] = []
    for attr, label in _SYNC_FIELDS:
        if attr not in values:
            continue
        old = getattr(existing, attr, None)
        new = values[attr]
        old_s = "" if old is None else str(old)
        new_s = "" if new is None else str(new)
        if old_s != new_s:
            out.append(JiraFieldChange(
                field=attr, label=label, old=old_s[:300], new=new_s[:300],
            ))
    return out


# 연결 해제 시 비워야 할 Jira 유래 컬럼 전부. 하나라도 빠뜨리면 Epic/컴포넌트 같은 잔재가
# 남아 "연결을 끊었는데 Jira 값이 보이는" 상태가 된다.
_JIRA_LINK_ATTRS: tuple[str, ...] = (
    "jira_issue_id", "jira_issue_key", "jira_url", "jira_status", "jira_status_category",
    "jira_updated_at", "jira_synced_at", "jira_epic", "jira_epic_key", "jira_epic_summary",
    "jira_issue_type", "jira_parent_key", "jira_parent_summary",
    "jira_components", "jira_labels",
)

_ISSUE_KEY_RE = re.compile(r"([A-Za-z][A-Za-z0-9_]*-\d+)")


def _clear_jira_link(item: WorkItem) -> None:
    """업무에서 Jira 연결 흔적을 모두 지운다 (Jira 쪽은 건드리지 않는다).

    `jira_issue_key` 가 비면 프로비저닝이 다시 열리므로, 잘못된 프로젝트에 만들어진 이슈를
    지우고 올바른 곳에 재생성하는 복구 경로가 성립한다."""
    for attr in _JIRA_LINK_ATTRS:
        setattr(item, attr, None)


def _parse_issue_key(raw: str) -> str:
    """`DL-42` · `https://jira/browse/DL-42?x=1` · 공백/소문자 입력에서 이슈 키를 뽑는다.

    사용자가 브라우저 주소창을 그대로 붙여넣는 경우가 대부분이라 URL 을 먼저 받아준다.
    키 형태가 아니면 빈 문자열 — 호출부가 거절한다."""
    text = (raw or "").strip()
    if not text:
        return ""
    # URL 이면 `/browse/<key>` 뒤쪽을 우선 본다(쿼리스트링에 다른 키가 섞여도 오인하지 않게).
    browse = re.search(r"/browse/([A-Za-z][A-Za-z0-9_]*-\d+)", text)
    if browse:
        return browse.group(1).upper()
    m = _ISSUE_KEY_RE.fullmatch(text) or _ISSUE_KEY_RE.search(text)
    return m.group(1).upper() if m else ""


def _apply_jira_fields(item: WorkItem, fields: dict, *, now: datetime) -> None:
    """`_jira_sync_values` 결과를 업무에 반영. 담당자/완료일은 비어 있을 때만 채운다."""
    for attr, val in _jira_sync_values(item, fields).items():
        setattr(item, attr, val)
    item.jira_synced_at = now
    if fields.get("closed_at") and not item.closed_at:
        item.closed_at = fields["closed_at"]
    if not (item.primary_assignee or "").strip() or item.primary_assignee == "(미할당)":
        item.primary_assignee = fields["primary_assignee"]
        item.assignee = fields["primary_assignee"]


# ── Confluence (Jira 와 같은 IdP 세션으로 연동) ─────────────────────────────────
@router.post("/confluence/test", response_model=JiraTestResult)
async def confluence_test(db: Session = Depends(get_db), actor: User = Depends(get_current_user)):
    """Confluence 연결 테스트 — 세션 만료 시 저장된 SSO 로그인으로 자동 재로그인 포함."""
    cfg = _get_config(db)
    svc, res = await _confluence_service_verified(db, actor, cfg)
    if svc is None or res.get("status") != "ok":
        return JiraTestResult(ok=False, detail=res.get("detail", "Confluence 연결 실패"))
    return JiraTestResult(ok=True, detail="연결 정상", display_name=res.get("display_name"))


@router.get("/confluence/search", response_model=ConfluenceSearchResult)
async def confluence_search(
    cql: str,
    limit: int = 25,
    db: Session = Depends(get_db),
    actor: User = Depends(get_current_user),
):
    """CQL 콘텐츠 검색 — SSO 로 캡처한 본인 세션(본인 권한)으로 실행."""
    cfg = _get_config(db)
    svc, res = await _confluence_service_verified(db, actor, cfg)
    if svc is None or res.get("status") != "ok":
        return ConfluenceSearchResult(status="error", detail=res.get("detail", "Confluence 세션 없음"))
    found = await svc.search(cql, limit=limit)
    return ConfluenceSearchResult(
        status=found.get("status", "error"),
        detail=found.get("detail", ""),
        total=found.get("total", 0),
        items=[ConfluenceSearchItem(**i) for i in found.get("items", [])],
    )


@router.post("/refresh/{item_id}", response_model=JiraImportResult)
async def refresh_work_item_from_jira(
    item_id: str,
    db: Session = Depends(get_db),
    actor: User = Depends(require_operator),
):
    """업무 1건을 연결된 Jira 이슈 기준으로 다시 가져온다 (게시판 행 단위 동기화).

    전체 가져오기와 동일한 매핑/보존 규칙을 쓰고, 변경 내역을 `items[0].changes` 로 돌려준다."""
    item = db.query(WorkItem).filter(WorkItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="업무를 찾을 수 없습니다.")
    if not item.jira_issue_key:
        return JiraImportResult(status="error", detail="Jira 와 연결되지 않은 업무입니다.")

    cfg = _get_config(db)
    base_url = cfg.get("base_url", "")
    if not base_url:
        return JiraImportResult(status="error", detail="Jira URL 미설정.")
    svc, _myself = await _jira_service_verified(db, actor, cfg)
    if svc is None:
        return JiraImportResult(status="error", detail="내 Jira 인증이 등록되지 않았습니다.")

    epic_field = (cfg.get("jira_epic_field") or "").strip()
    got = await svc.get_issue(item.jira_issue_key,
                              fields=ISSUE_FIELDS + ([epic_field] if epic_field else []))
    if got.get("missing"):
        # 삭제됐거나 내 권한으로 안 보이거나 — 서버는 구분할 수 없다. 연결을 자동으로
        # 끊지 않고 상태만 알려, 화면에서 사용자가 해제/삭제를 고르게 한다.
        return JiraImportResult(
            status="missing", total=1,
            detail=(f"Jira 에서 {item.jira_issue_key} 를 찾을 수 없습니다 — "
                    "삭제됐거나 조회 권한이 없습니다. 연결을 해제하거나 다른 이슈로 바꿀 수 있습니다."),
        )
    if got.get("status") != "ok":
        return JiraImportResult(status=got.get("status", "error"),
                                detail=got.get("detail", "Jira 이슈 조회 실패"))
    confluence_base = (cfg.get("confluence_base_url") or "").strip()
    fields = map_jira_issue(got["issue"], base_url,
                            assignee_resolver=_build_assignee_resolver(db), epic_field=epic_field,
                            confluence_base_url=confluence_base)
    # 행 단위라 이슈 1건 — 본문에 없더라도 원격 링크에 붙은 Confluence 문서를 찾아본다
    # (대량 가져오기는 이슈마다 1콜이 되어 N+1 이므로 본문 스캔만 한다).
    if confluence_base and not fields.get("confluence_url"):
        linked = await svc.remote_links(item.jira_issue_key)
        for link in linked.get("links", []):
            if link["url"].rstrip("/").startswith(confluence_base.rstrip("/")):
                fields["confluence_url"] = link["url"][:500]
                break
    changes = _diff_existing(item, fields)
    if not changes:
        item.jira_synced_at = datetime.utcnow()
        db.commit()
        return JiraImportResult(
            status="ok", detail="변경 사항이 없습니다.", total=1, skipped=1,
            items=[JiraImportItemPreview(
                jira_key=item.jira_issue_key, title=fields["title"],
                kanban_status=fields["kanban_status"], action="unchanged")],
        )

    _apply_jira_fields(item, fields, now=datetime.utcnow())
    db.commit()
    audit_logger.record(
        db, action="work_item.jira_refresh", actor=actor,
        target_type="work_item", target_id=str(item.id),
        details={"jira_key": item.jira_issue_key, "changed": [c.field for c in changes]},
    )
    return JiraImportResult(
        status="ok", detail=f"{len(changes)}개 필드가 갱신되었습니다.", total=1, updated=1,
        items=[JiraImportItemPreview(
            jira_key=item.jira_issue_key, title=fields["title"],
            kanban_status=fields["kanban_status"], action="update", changes=changes)],
    )


# ── PEP → Jira 신규 생성 / 삭제 ────────────────────────────────────────────────
@router.post("/create", response_model=JiraCreateResult)
async def create_jira_issue(
    payload: JiraCreateRequest,
    db: Session = Depends(get_db),
    actor: User = Depends(require_operator),
):
    """PEP 에서 Jira 이슈를 새로 만든다. work_item_id 를 주면 그 업무 내용으로 만들고
    생성된 키/URL 을 업무에 연결해 이후 push/가져오기가 이어지게 한다."""
    cfg = _get_config(db)
    if not cfg.get("base_url") or not cfg.get("enabled", False):
        return JiraCreateResult(status="error", detail="Jira 연동이 비활성화되었거나 URL 미설정.")

    item: Optional[WorkItem] = None
    if payload.work_item_id:
        item = db.query(WorkItem).filter(WorkItem.id == payload.work_item_id).first()
        if not item:
            raise HTTPException(status_code=404, detail="업무를 찾을 수 없습니다.")
        if item.jira_issue_key:
            return JiraCreateResult(status="error",
                                    detail=f"이미 Jira({item.jira_issue_key})와 연결된 업무입니다.")

    project_key = (payload.project_key or cfg.get("default_project_key") or "").strip()
    if not project_key:
        return JiraCreateResult(status="error", detail="프로젝트 키를 지정하세요 (또는 공통 설정의 기본 프로젝트).")
    summary = (payload.summary or (item.title if item else "") or "").strip()
    if not summary:
        return JiraCreateResult(status="error", detail="제목(summary)을 입력하세요.")
    description = payload.description if payload.description is not None else (item.content if item else "")
    priority = payload.priority or (PEP_PRIORITY_TO_JIRA.get(item.priority) if item else None)

    svc, _myself = await _jira_service_verified(db, actor, cfg)
    if svc is None:
        return JiraCreateResult(status="error", detail="내 Jira 인증이 등록되지 않았습니다 (설정 > 연동).")
    res = await svc.create_issue(
        project_key, summary, description=description or "", issue_type=payload.issue_type,
        priority=priority, labels=payload.labels, components=payload.components,
    )
    if res.get("status") != "ok":
        return JiraCreateResult(status=res.get("status", "error"),
                                detail=res.get("detail", "이슈 생성 실패"))

    linked_id = None
    if item:
        item.jira_issue_key = res.get("key")
        item.jira_url = res.get("url")
        item.jira_issue_id = res.get("id") or None
        item.jira_synced_at = datetime.utcnow()
        db.commit()
        linked_id = str(item.id)
    audit_logger.record(
        db, action="work_item.jira_create", actor=actor,
        target_type="jira_issue", target_id=res.get("key"),
        details={"project": project_key, "work_item_id": linked_id},
    )
    return JiraCreateResult(status="ok", detail="Jira 이슈가 생성되었습니다.",
                            jira_key=res.get("key"), jira_url=res.get("url"),
                            linked_work_item_id=linked_id)


@router.delete("/issue/{key}", response_model=JiraDeleteResult)
async def delete_jira_issue(
    key: str,
    db: Session = Depends(get_db),
    actor: User = Depends(require_operator),
):
    """Jira 이슈 삭제 — 연결된 PEP 업무가 있으면 연결만 해제한다(업무는 보존)."""
    cfg = _get_config(db)
    if not cfg.get("base_url"):
        return JiraDeleteResult(status="error", detail="Jira URL 미설정.")
    svc, _myself = await _jira_service_verified(db, actor, cfg)
    if svc is None:
        return JiraDeleteResult(status="error", detail="내 Jira 인증이 등록되지 않았습니다.")
    res = await svc.delete_issue(key)
    if res.get("status") != "ok":
        return JiraDeleteResult(status=res.get("status", "error"),
                                detail=res.get("detail", "이슈 삭제 실패"))
    unlinked = None
    item = db.query(WorkItem).filter(WorkItem.jira_issue_key == key).first()
    if item:
        _clear_jira_link(item)
        db.commit()
        unlinked = str(item.id)
    audit_logger.record(
        db, action="work_item.jira_delete", actor=actor,
        target_type="jira_issue", target_id=key, details={"unlinked_work_item_id": unlinked},
    )
    return JiraDeleteResult(status="ok", detail=f"Jira {key} 삭제됨", unlinked_work_item_id=unlinked)


# ── 연결 복구 (Jira 쪽은 건드리지 않고 PEP 연결만 정리/교체) ─────────────────────
@router.post("/unlink/{item_id}", response_model=JiraUnlinkResult)
def unlink_work_item(
    item_id: str,
    payload: JiraUnlinkRequest,
    db: Session = Depends(get_db),
    actor: User = Depends(require_operator),
):
    """업무의 Jira 연결을 끊는다 (선택적으로 업무 행까지 삭제).

    Jira 에서 이슈를 **이미 직접 지운** 뒤 PEP 에 남은 죽은 링크를 정리하는 경로다.
    `DELETE /jira/issue/{key}` 는 Jira 에서 먼저 지우는 흐름이라, 이미 없는 이슈에는
    Jira 가 404 를 돌려줘 해제까지 도달하지 못한다 — 그래서 별도 엔드포인트가 필요하다.

    연결을 끊으면 `jira_issue_key` 가 비어 프로비저닝이 다시 열리므로, 잘못된 프로젝트에
    만들어진 이슈를 지우고 올바른 곳에 재생성하는 복구가 가능해진다."""
    item = db.query(WorkItem).filter(WorkItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="업무를 찾을 수 없습니다.")

    prev_key = item.jira_issue_key or ""
    if payload.delete_work_item:
        # 업무 삭제 권한은 업무 관리와 동일 규칙(등록자/담당자/admin)을 그대로 쓴다.
        _assert_work_item_ownership(item, actor, op="삭제", db=db)
        wid = str(item.id)
        db.delete(item)
        db.commit()
        audit_logger.record(
            db, action="work_item.jira_unlink", actor=actor,
            target_type="work_item", target_id=wid,
            details={"jira_key": prev_key or None, "deleted": True},
        )
        return JiraUnlinkResult(
            status="ok", work_item_id=wid, work_item_deleted=True,
            detail=(f"{prev_key} 연결을 끊고 업무를 삭제했습니다." if prev_key
                    else "업무를 삭제했습니다."),
        )

    if not prev_key:
        return JiraUnlinkResult(status="ok", work_item_id=str(item.id),
                                detail="이미 Jira 와 연결돼 있지 않습니다.")
    _clear_jira_link(item)
    db.commit()
    audit_logger.record(
        db, action="work_item.jira_unlink", actor=actor,
        target_type="work_item", target_id=str(item.id),
        details={"jira_key": prev_key, "deleted": False},
    )
    return JiraUnlinkResult(
        status="ok", work_item_id=str(item.id),
        detail=f"{prev_key} 연결을 해제했습니다 — 이제 Jira·Confluence 자동 생성을 다시 할 수 있습니다.",
    )


@router.post("/relink/{item_id}", response_model=JiraRelinkResult)
async def relink_work_item(
    item_id: str,
    payload: JiraRelinkRequest,
    db: Session = Depends(get_db),
    actor: User = Depends(require_operator),
):
    """연결을 다른 Jira 이슈로 갈아끼운다 (이슈 키 또는 브라우저 URL 입력).

    **Jira 에서 실제로 조회해 존재를 확인한 뒤에만** 연결한다 — 검증 없이 키를 받으면
    또 다른 죽은 링크가 생기고, 그게 애초에 이 기능이 필요해진 이유다."""
    item = db.query(WorkItem).filter(WorkItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="업무를 찾을 수 없습니다.")

    key = _parse_issue_key(payload.key_or_url)
    if not key:
        return JiraRelinkResult(status="error",
                                detail="이슈 키를 찾을 수 없습니다 (예: DL-42 또는 .../browse/DL-42).")

    dup = (
        db.query(WorkItem)
        .filter(WorkItem.jira_issue_key == key, WorkItem.id != item.id)
        .first()
    )
    if dup:
        return JiraRelinkResult(
            status="error",
            detail=f"{key} 는 이미 다른 업무({dup.title or dup.category})에 연결돼 있습니다.",
        )

    cfg = _get_config(db)
    base_url = cfg.get("base_url", "")
    if not base_url:
        return JiraRelinkResult(status="error", detail="Jira URL 미설정.")
    svc, _myself = await _jira_service_verified(db, actor, cfg)
    if svc is None:
        return JiraRelinkResult(status="error", detail="내 Jira 인증이 등록되지 않았습니다.")

    epic_field = (cfg.get("jira_epic_field") or "").strip()
    got = await svc.get_issue(key, fields=ISSUE_FIELDS + ([epic_field] if epic_field else []))
    if got.get("missing"):
        return JiraRelinkResult(
            status="missing",
            detail=f"Jira 에서 {key} 를 찾을 수 없습니다 (삭제됐거나 조회 권한이 없습니다).",
        )
    if got.get("status") != "ok":
        return JiraRelinkResult(status=got.get("status", "error"),
                                detail=got.get("detail", "Jira 이슈 조회 실패"))

    prev_key = item.jira_issue_key or ""
    # 이전 연결의 잔재(Epic/컴포넌트 등)를 먼저 비우고 새 이슈 값으로 채운다.
    _clear_jira_link(item)
    fields = map_jira_issue(
        got["issue"], base_url, assignee_resolver=_build_assignee_resolver(db),
        epic_field=epic_field, confluence_base_url=(cfg.get("confluence_base_url") or "").strip(),
    )
    item.jira_issue_id = fields.get("jira_issue_id") or None
    _apply_jira_fields(item, fields, now=datetime.utcnow())
    db.commit()
    audit_logger.record(
        db, action="work_item.jira_relink", actor=actor,
        target_type="work_item", target_id=str(item.id),
        details={"from": prev_key or None, "to": key},
    )
    return JiraRelinkResult(status="ok", jira_key=key, jira_url=item.jira_url,
                            detail=f"{key} 로 연결했습니다.")


# 고아 점검 상한 — 키마다 1콜이라 무한정 돌지 않게 자른다(초과분은 truncated 로 알림).
_VERIFY_LINKS_MAX = 200


@router.post("/verify-links", response_model=JiraVerifyLinksResult)
async def verify_links(
    all_users: bool = False,
    db: Session = Depends(get_db),
    actor: User = Depends(require_operator),
):
    """연결된 업무들의 Jira 이슈가 아직 살아 있는지 확인해 **죽은 링크 목록**을 돌려준다.

    확인은 키마다 `GET /issue/{key}` 개별 호출로 한다. `issuekey in (...)` 벌크 JQL 은
    존재하지 않는 키가 하나만 섞여도 Jira 가 쿼리 전체를 400 으로 거절해서 — 정확히
    우리가 찾으려는 그 상황에서 — 쓸 수 없다.

    삭제와 권한없음을 구분할 수 없으므로 여기서 정리하지 않는다. 사용자가 목록에서
    골라 `POST /jira/unlink/{item_id}` 로 처리한다."""
    cfg = _get_config(db)
    if not cfg.get("base_url"):
        return JiraVerifyLinksResult(status="error", detail="Jira URL 미설정.")
    svc, _myself = await _jira_service_verified(db, actor, cfg)
    if svc is None:
        return JiraVerifyLinksResult(status="error", detail="내 Jira 인증이 등록되지 않았습니다.")

    q = db.query(WorkItem).filter(WorkItem.jira_issue_key.isnot(None))
    if not all_users:
        # 기본은 내가 담당이거나 내가 가져온(watcher) 업무만 — 남의 업무까지 훑지 않는다.
        name = (actor.display_name or actor.username or "").strip()
        q = q.filter(
            WorkItem.jira_watchers.contains([actor.username])
            | WorkItem.primary_assignee.ilike(f"%{name}%")
            | WorkItem.secondary_assignee.ilike(f"%{name}%")
        )
    rows = q.order_by(WorkItem.updated_at.desc()).limit(_VERIFY_LINKS_MAX + 1).all()
    truncated = len(rows) > _VERIFY_LINKS_MAX
    rows = rows[:_VERIFY_LINKS_MAX]
    if not rows:
        return JiraVerifyLinksResult(status="ok", detail="확인할 Jira 연결이 없습니다.")

    sem = asyncio.Semaphore(5)

    async def _check(item: WorkItem):
        async with sem:
            return item, await svc.get_issue(item.jira_issue_key, fields=["summary"])

    missing: list[JiraMissingLink] = []
    for item, res in await asyncio.gather(*(_check(r) for r in rows)):
        if res.get("missing"):
            missing.append(JiraMissingLink(
                work_item_id=str(item.id), jira_key=item.jira_issue_key or "",
                title=item.title or item.category or "",
                detail=res.get("detail", ""),
            ))

    detail = (f"{len(rows)}건 중 {len(missing)}건이 Jira 에서 확인되지 않습니다."
              if missing else f"{len(rows)}건 모두 정상입니다.")
    return JiraVerifyLinksResult(status="ok", detail=detail, checked=len(rows),
                                 missing=missing, truncated=truncated)


# ── 업무 등록 시 Jira + Confluence 동시 생성 (프로비저닝) ────────────────────────
# 사용자별 "기준 조건" 프리셋 (user_settings). 매 등록마다 프로젝트/컴포넌트/라벨/Epic/
# 스페이스를 다시 입력하지 않도록, 마지막에 성공한 조건을 이 키에 저장해 다음 등록의
# 기본값으로 쓴다. 관리자 공통 설정보다 우선하되 화면에서 언제든 수정 가능하다.
PROVISION_PRESET_KEY = "jira_provision_preset"
_PRESET_FIELDS = (
    "project_key", "issue_type", "priority", "labels", "components",
    "epic_key", "parent_key", "space_key", "parent_page_id",
)


def _load_provision_preset(db: Session, user_id: str) -> dict:
    """저장된 프리셋 — 형식이 깨진 값은 무시하고 빈 dict 로 폴백(개인 설정은 best-effort)."""
    raw = get_user_setting(db, user_id, PROVISION_PRESET_KEY, {}) or {}
    if not isinstance(raw, dict):
        return {}
    out: dict = {}
    for key in _PRESET_FIELDS:
        val = raw.get(key)
        if key in ("labels", "components"):
            if isinstance(val, list):
                cleaned = [str(v).strip() for v in val if str(v).strip()]
                if cleaned:
                    out[key] = cleaned
        elif isinstance(val, str) and val.strip():
            out[key] = val.strip()
    return out


def _save_provision_preset(db: Session, user_id: str, payload: "ProvisionRequest") -> None:
    """이번에 쓴 조건을 프리셋으로 저장. 저장 실패가 생성 결과를 뒤집지 않도록 예외를 삼킨다."""
    preset = {
        "project_key": (payload.project_key or "").strip(),
        "issue_type": (payload.issue_type or "").strip(),
        "priority": (payload.priority or "").strip(),
        "labels": [x.strip() for x in payload.labels if x.strip()],
        "components": [x.strip() for x in payload.components if x.strip()],
        "epic_key": (payload.epic_key or "").strip(),
        "parent_key": (payload.parent_key or "").strip(),
        "space_key": (payload.space_key or "").strip(),
        "parent_page_id": (payload.parent_page_id or "").strip(),
    }
    try:
        set_user_setting(db, user_id, PROVISION_PRESET_KEY, preset)
    except Exception as exc:  # noqa: BLE001 — 개인 기본값 저장은 부가 기능
        db.rollback()
        logger.warning("프로비저닝 프리셋 저장 실패 (%s): %s", user_id, exc)


def _default_page_body(item: WorkItem, jira_key: str = "", jira_url: str = "") -> str:
    """업무 내용을 담은 기본 Confluence 문서(storage format).

    사용자가 본문을 따로 주지 않으면 이 골격으로 만든다 — 담당자/일정/Jira 링크가 들어간
    한 장짜리 작업 문서."""
    rows = [
        ("담당자", (item.primary_assignee or item.assignee or "")),
        ("시작일", item.started_at.strftime("%Y-%m-%d") if item.started_at else ""),
        ("구분", item.category or ""),
        ("우선순위", item.priority or ""),
    ]
    if jira_key:
        link = f'<a href="{html.escape(jira_url)}">{html.escape(jira_key)}</a>' if jira_url else html.escape(jira_key)
        rows.append(("Jira", link))
    meta = "".join(
        f"<tr><th>{html.escape(k)}</th><td>{v if k == 'Jira' else html.escape(str(v))}</td></tr>"
        for k, v in rows
    )
    content = html.escape((item.content or "").strip()) or "(내용 없음)"
    return (
        f"<table><tbody>{meta}</tbody></table>"
        f"<h2>작업 내용</h2><p>{content}</p>"
        "<h2>진행 경과</h2><p></p>"
        "<h2>이슈 / 리스크</h2><p></p>"
        "<h2>결과 / 후속 조치</h2><p></p>"
    )


@router.get("/provision/defaults", response_model=ProvisionDefaults)
async def provision_defaults(
    work_item_id: Optional[str] = None,
    db: Session = Depends(get_db),
    actor: User = Depends(get_current_user),
):
    """업무 → Jira/Confluence 생성 화면의 **기본값**.

    관리자 공통 설정 + 주간보고 저장 위치 + 사용자의 업무 내용에서 최대한 채워 돌려주고,
    화면에서는 그대로 수정 가능하게 한다(강제하지 않는다)."""
    cfg = _get_config(db)
    weekly = _get_weekly_settings(db)
    item = db.query(WorkItem).filter(WorkItem.id == work_item_id).first() if work_item_id else None

    cred = db.query(UserJiraCredential).filter(UserJiraCredential.username == actor.username).first()
    missing: list[str] = []
    if not cfg.get("base_url"):
        missing.append("Jira URL(관리자 설정)")
    if not cred:
        missing.append("내 Jira 인증")
    if not (cfg.get("confluence_base_url") or "").strip():
        missing.append("Confluence URL(관리자 설정)")

    title = (item.title if item else "") or ""
    # 공통 설정/업무 내용에서 만든 기본값 위에 **내 프리셋**을 덮어쓴다 —
    # "처음 입력한 조건을 다음부터 재사용" 이 여기서 성립한다.
    base = {
        "project_key": (cfg.get("default_project_key") or ""),
        "issue_type": "Task",
        "priority": PEP_PRIORITY_TO_JIRA.get((item.priority if item else "") or "", ""),
        "labels": [],
        "components": [c for c in [(item.category if item else "")] if c],
        "epic_key": "",
        "parent_key": "",
        "space_key": (weekly.get("space_key") or ""),
        "parent_page_id": (weekly.get("parent_page_id") or ""),
    }
    preset = _load_provision_preset(db, actor.id)
    base.update(preset)
    # 가져온 업무라면 Jira 가 알려준 실제 계층/컴포넌트가 프리셋보다 정확하다.
    if item is not None:
        if item.jira_epic_key:
            base["epic_key"] = item.jira_epic_key
        if item.jira_components:
            base["components"] = list(item.jira_components)

    return ProvisionDefaults(
        jira_enabled=bool(cfg.get("base_url") and cfg.get("enabled", False) and cred),
        confluence_enabled=bool((cfg.get("confluence_base_url") or "").strip() and cred),
        summary=title,
        description=(item.content if item else "") or "",
        page_title=title,
        reporter=(cred.jira_account if cred and cred.jira_account else actor.username),
        preset_source="user" if preset else "settings",
        detail=("바로 생성할 수 있습니다." if not missing
                else "미설정: " + ", ".join(missing)),
        **base,
    )


@router.post("/provision", response_model=ProvisionResult)
async def provision_work_item(
    payload: ProvisionRequest,
    db: Session = Depends(get_db),
    actor: User = Depends(require_operator),
):
    """업무 1건을 Jira 이슈 + Confluence 문서로 만들고 둘 다 업무에 연결한다.

    한쪽이 실패해도 다른 쪽 결과는 유지하고 `partial` 로 알린다(둘 다 실패면 error).
    Confluence 문서에는 생성된 Jira 키 링크가 함께 들어간다."""
    item = db.query(WorkItem).filter(WorkItem.id == payload.work_item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="업무를 찾을 수 없습니다.")
    cfg = _get_config(db)

    jira_key = jira_url = ""
    jira_detail = confluence_detail = ""
    conf_id = conf_url = ""
    jira_ok = conf_ok = False
    # 실패가 "내 인증(토큰/세션)" 문제인지 — 프론트가 재시도 전에 연결 설정 카드를
    # 보여줄지 판단하는 신호. 빈 필드 같은 입력값 문제와는 구분한다.
    jira_auth_issue = confluence_auth_issue = False

    # ── Jira ─────────────────────────────────────────────────────────────────
    if payload.create_jira:
        if item.jira_issue_key:
            jira_detail = (f"이미 {item.jira_issue_key} 와 연결돼 있어 생성을 건너뛰었습니다 — "
                           "다른 프로젝트로 다시 만들려면 먼저 Jira 연결을 해제하세요.")
            jira_key, jira_url = item.jira_issue_key, item.jira_url or ""
            jira_ok = True
        elif not cfg.get("base_url"):
            jira_detail = "Jira URL 미설정."
        else:
            svc, _myself = await _jira_service_verified(db, actor, cfg)
            project_key = (payload.project_key or cfg.get("default_project_key") or "").strip()
            summary = (payload.summary or item.title or "").strip()
            if svc is None:
                jira_detail = "내 Jira 인증이 등록되지 않았습니다."
                jira_auth_issue = True
            elif not project_key:
                jira_detail = "프로젝트 키를 지정하세요."
            elif not summary:
                jira_detail = "제목(summary)이 비어 있습니다."
            else:
                epic_key = (payload.epic_key or "").strip()
                parent_key = (payload.parent_key or "").strip()
                res = await svc.create_issue(
                    project_key, summary,
                    description=(payload.description if payload.description is not None else item.content) or "",
                    issue_type=payload.issue_type,
                    priority=payload.priority or PEP_PRIORITY_TO_JIRA.get(item.priority or ""),
                    labels=payload.labels, components=payload.components,
                    epic_key=epic_key, epic_field=(cfg.get("jira_epic_field") or "").strip(),
                    parent_key=parent_key,
                )
                if res.get("status") == "ok":
                    jira_ok = True
                    jira_key, jira_url = res.get("key", ""), res.get("url", "")
                    item.jira_issue_key = jira_key
                    item.jira_issue_id = res.get("id") or None
                    item.jira_url = jira_url
                    item.jira_issue_type = payload.issue_type or None
                    item.jira_components = payload.components or None
                    item.jira_labels = payload.labels or None
                    if epic_key:
                        item.jira_epic_key = epic_key
                        item.jira_epic = (item.jira_epic or epic_key)
                    if parent_key:
                        item.jira_parent_key = parent_key
                    item.jira_synced_at = datetime.utcnow()
                else:
                    jira_detail = res.get("detail", "Jira 이슈 생성 실패")
                    jira_auth_issue = bool(res.get("auth_failed"))

    # ── Confluence ───────────────────────────────────────────────────────────
    if payload.create_confluence:
        if item.confluence_page_id:
            # Jira 와 동일하게 이미 연결된 쪽은 건너뛴다 — 나머지 한쪽만 재시도하는
            # 호출(둘 다 True 로 다시 보내도)이 이미 성공한 페이지를 불필요하게
            # 새 버전으로 갱신하지 않도록 한다.
            confluence_detail = "이미 연결된 Confluence 문서가 있어 생성을 건너뛰었습니다."
            conf_id, conf_url = item.confluence_page_id, item.confluence_url or ""
            conf_ok = True
        else:
            space_key = (payload.space_key or "").strip() or (_get_weekly_settings(db).get("space_key") or "").strip()
            page_title = (payload.page_title or item.title or "").strip()
            if not space_key:
                confluence_detail = "Confluence 스페이스 키를 지정하세요."
            elif not page_title:
                confluence_detail = "문서 제목이 비어 있습니다."
            else:
                svc, res = await _confluence_service_verified(db, actor, cfg)
                if svc is None or res.get("status") != "ok":
                    confluence_detail = res.get("detail", "Confluence 세션 없음")
                    confluence_auth_issue = True
                else:
                    body = payload.page_body or _default_page_body(item, jira_key, jira_url)
                    out = await svc.upsert_page(
                        space_key, page_title, body,
                        parent_id=(payload.parent_page_id or "").strip()
                        or (_get_weekly_settings(db).get("parent_page_id") or ""),
                    )
                    if out.get("status") == "ok":
                        conf_ok = True
                        conf_id, conf_url = out.get("id", ""), out.get("url", "")
                        item.confluence_page_id = conf_id or None
                        item.confluence_url = conf_url or None
                    else:
                        confluence_detail = out.get("detail", "Confluence 문서 생성 실패")
                        confluence_auth_issue = bool(out.get("auth_failed"))

    wanted = [payload.create_jira, payload.create_confluence]
    succeeded = [payload.create_jira and jira_ok, payload.create_confluence and conf_ok]
    if not any(wanted):
        return ProvisionResult(status="error", detail="생성할 대상을 하나 이상 선택하세요.")
    if all(s for s, w in zip(succeeded, wanted) if w):
        status, detail = "ok", "생성이 완료되었습니다."
    elif any(succeeded):
        status, detail = "partial", "일부만 생성되었습니다 — 아래 사유를 확인하세요."
    else:
        status, detail = "error", "생성에 실패했습니다 — 아래 사유를 확인하세요."

    # 다음 시도(재시도 버튼/게시판 재반영)가 무엇이 왜 막혔는지 알 수 있도록 결과를
    # 업무에 영속화한다 — null 이면 프로비저닝을 시도한 적 없다는 뜻(가져오기/수동
    # 등록과 구분).
    item.provision_status = status
    item.provision_jira_error = jira_detail if (payload.create_jira and not jira_ok) else None
    item.provision_confluence_error = confluence_detail if (payload.create_confluence and not conf_ok) else None
    db.commit()

    audit_logger.record(
        db, action="work_item.provision", actor=actor,
        target_type="work_item", target_id=str(item.id),
        details={"jira": jira_key or None, "confluence": conf_id or None},
    )
    # 하나라도 성공했으면 이번 조건을 내 기본값으로 기억한다(다음 등록에서 자동 채움).
    if payload.remember_preset and (jira_ok or conf_ok):
        _save_provision_preset(db, actor.id, payload)

    return ProvisionResult(
        status=status, detail=detail,
        jira_key=jira_key or None, jira_url=jira_url or None, jira_detail=jira_detail,
        confluence_page_id=conf_id or None, confluence_url=conf_url or None,
        confluence_detail=confluence_detail,
        jira_auth_issue=jira_auth_issue, confluence_auth_issue=confluence_auth_issue,
    )


# ── 주간보고 (월~금 집계 → 3개 표 → Confluence 게시) ────────────────────────────
WEEKLY_SETTINGS_KEY = "jira_weekly_report"
DEFAULT_WEEKLY_SETTINGS = {
    # Jira WBS/간트 차트 링크 — 진척률 표 위에 노출한다(플러그인/보드마다 URL 이 달라 설정값).
    "gantt_url": "",
    "space_key": "",
    "parent_page_id": "",
    "title_template": "주간보고 {start} ~ {end}",
    "auto_enabled": False,
    "auto_cron": "0 17 * * 5",
    "project_filter": "",
}


def _get_weekly_settings(db: Session) -> dict:
    row = db.query(AppSetting).filter(AppSetting.key == WEEKLY_SETTINGS_KEY).first()
    value = dict(DEFAULT_WEEKLY_SETTINGS)
    if row and isinstance(row.value, dict):
        value.update(row.value)
    return value


def _parse_week_of(value: Optional[str]):
    if not value:
        return None
    try:
        return datetime.strptime(value.strip(), "%Y-%m-%d").date()
    except ValueError:
        return None


@router.get("/weekly-report/settings", response_model=WeeklyReportSettings)
def get_weekly_settings(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return WeeklyReportSettings(**_get_weekly_settings(db))


@router.put("/weekly-report/settings", response_model=WeeklyReportSettings)
def update_weekly_settings(
    payload: WeeklyReportSettings,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    row = db.query(AppSetting).filter(AppSetting.key == WEEKLY_SETTINGS_KEY).first()
    value = payload.model_dump()
    if row:
        row.value = value
    else:
        db.add(AppSetting(key=WEEKLY_SETTINGS_KEY, value=value))
    db.commit()
    return WeeklyReportSettings(**value)


@router.post("/weekly-report/preview", response_model=WeeklyReport)
def weekly_report_preview(
    payload: Optional[WeeklyReportRequest] = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """주간보고 미리보기 — 게시하지 않고 표 데이터만 만든다."""
    payload = payload or WeeklyReportRequest()
    settings = _get_weekly_settings(db)
    report = weekly_report_service.build_report(
        db, anchor=_parse_week_of(payload.week_of),
        project_filter=payload.project_filter or settings.get("project_filter", ""),
        base_url=_get_config(db).get("base_url", ""),
    )
    return WeeklyReport(**report)


@router.post("/weekly-report/publish", response_model=WeeklyPublishResult)
async def weekly_report_publish(
    payload: Optional[WeeklyPublishRequest] = None,
    db: Session = Depends(get_db),
    actor: User = Depends(require_operator),
):
    """주간보고를 만들어 Confluence 에 게시한다(같은 제목이면 새 버전으로 갱신).

    저장 위치(스페이스/부모/제목)는 요청에서 덮어쓸 수 있고, 미지정 시 설정값을 쓴다."""
    payload = payload or WeeklyPublishRequest()
    cfg = _get_config(db)
    settings = _get_weekly_settings(db)
    report = weekly_report_service.build_report(
        db, anchor=_parse_week_of(payload.week_of),
        project_filter=payload.project_filter or settings.get("project_filter", ""),
        base_url=cfg.get("base_url", ""),
    )
    space_key = (payload.space_key or settings.get("space_key") or "").strip()
    if not space_key:
        return WeeklyPublishResult(status="error",
                                   detail="Confluence 스페이스 키를 지정하세요 (주간보고 설정).")
    title = (payload.title or "").strip() or (
        settings.get("title_template") or "주간보고 {start} ~ {end}"
    ).format(start=report["period_start"], end=report["period_end"])

    svc, res = await _confluence_service_verified(db, actor, cfg)
    if svc is None or res.get("status") != "ok":
        return WeeklyPublishResult(status="error", detail=res.get("detail", "Confluence 세션 없음"))
    body = weekly_report_service.render_storage_html(report)
    out = await svc.upsert_page(
        space_key, title, body,
        parent_id=(payload.parent_page_id or settings.get("parent_page_id") or ""),
    )
    if out.get("status") != "ok":
        return WeeklyPublishResult(status=out.get("status", "error"),
                                   detail=out.get("detail", "Confluence 게시 실패"))
    audit_logger.record(
        db, action="work_item.weekly_report_publish", actor=actor,
        target_type="confluence_page", target_id=out.get("id"),
        details={"space": space_key, "title": title, "action": out.get("action")},
    )
    return WeeklyPublishResult(
        status="ok", detail=f"Confluence 에 {('생성' if out.get('action') == 'created' else '갱신')}되었습니다.",
        action=out.get("action", ""), page_url=out.get("url"), page_id=out.get("id"),
    )


# ── 가져오기 (단방향, upsert by jira_issue_id) ──────────────────────────────────
@router.post("/import", response_model=JiraImportResult)
async def import_issues(
    payload: JiraImportRequest,
    db: Session = Depends(get_db),
    actor: User = Depends(require_operator),
):
    cfg = _get_config(db)
    base_url = cfg.get("base_url", "")
    if not base_url or not cfg.get("enabled", False):
        return JiraImportResult(status="error", detail="Jira 연동이 비활성화되었거나 URL 미설정 (설정에서 활성화하세요).")
    token, _auth_type = _user_credential(db, actor.username)
    if not token:
        return JiraImportResult(status="error", detail="내 Jira 인증(PAT 또는 세션 쿠키)이 등록되지 않았습니다 (설정 > 연동에서 등록).")

    # JQL 구성
    if payload.scope == "me":
        jql = "assignee = currentUser() ORDER BY updated DESC"
    elif payload.scope == "project":
        pk = (payload.project_key or "").strip()
        if not pk:
            return JiraImportResult(status="error", detail="프로젝트 키를 입력하세요.")
        jql = f'project = "{_jql_quote(pk)}" ORDER BY updated DESC'
    elif payload.scope == "filter":
        jql, err = _build_filter_jql(payload)
        if err:
            return JiraImportResult(status="error", detail=err)
    else:  # jql
        jql = (payload.jql or "").strip()
        if not jql:
            return JiraImportResult(status="error", detail="JQL 을 입력하세요.")

    # 세션 만료(401) 시 저장된 SSO 로그인 정보로 자동 재로그인 포함.
    svc, _myself = await _jira_service_verified(db, actor, cfg)
    if svc is None:
        return JiraImportResult(status="error", detail="내 Jira 인증(PAT 또는 세션 쿠키)이 등록되지 않았습니다 (설정 > 연동에서 등록).")
    epic_field = (cfg.get("jira_epic_field") or "").strip()
    search = await svc.search(jql, extra_fields=[epic_field] if epic_field else None)
    if search.get("status") != "ok":
        return JiraImportResult(
            status=search.get("status", "error"),
            detail=search.get("detail", "Jira 검색 실패"),
            total=search.get("total", 0),
            applied_jql=jql,
        )

    resolver = _build_assignee_resolver(db)
    confluence_base = (cfg.get("confluence_base_url") or "").strip()
    issues = search.get("issues", [])
    created = updated = skipped = 0
    errors: list[str] = []
    preview: list[JiraImportItemPreview] = []
    now = datetime.utcnow()

    for issue in issues:
        try:
            fields = map_jira_issue(issue, base_url, assignee_resolver=resolver,
                                    epic_field=epic_field, confluence_base_url=confluence_base)
            jid = fields.get("jira_issue_id")
            if not jid:
                skipped += 1
                continue
            existing = db.query(WorkItem).filter(WorkItem.jira_issue_id == jid).first()
            if existing is None and fields.get("jira_issue_key"):
                # id 가 없던(Excel 로 들어왔거나 잘못 저장된) 기존 행은 **키(DL-#) 기준**으로
                # 찾아 덮어쓴다 — 중복 생성 대신 정정이 되게 한다.
                existing = (
                    db.query(WorkItem)
                    .filter(WorkItem.jira_issue_key == fields["jira_issue_key"])
                    .first()
                )
                if existing is not None:
                    existing.jira_issue_id = jid
            changes = _diff_existing(existing, fields) if existing else []
            if not existing:
                action = "create"
            elif changes:
                action = "update"
            else:
                action = "unchanged"
            preview.append(JiraImportItemPreview(
                jira_key=fields["jira_issue_key"], title=fields["title"],
                kanban_status=fields["kanban_status"], action=action, changes=changes,
            ))
            if payload.dry_run:
                if action == "create":
                    created += 1
                elif action == "update":
                    updated += 1
                else:
                    skipped += 1
                continue

            # 사용자가 미리보기에서 고른 항목만 적용 (비우면 전체).
            if payload.only_keys and fields["jira_issue_key"] not in payload.only_keys:
                skipped += 1
                continue
            if action == "unchanged":
                # 변경 없음 — 동기화 시각만 갱신하고 넘어간다.
                existing.jira_synced_at = now
                skipped += 1
                continue

            if existing:
                # Jira-소유 필드만 갱신 (PEP 로컬 편집 보존). 담당자는 비어있을 때만 채움.
                _apply_jira_fields(existing, fields, now=now)
                watchers = list(existing.jira_watchers or [])
                if actor.username not in watchers:
                    watchers.append(actor.username)
                existing.jira_watchers = watchers
                updated += 1
            else:
                item = WorkItem(
                    type=fields["type"],
                    type_label=fields["type_label"],
                    assignee=fields["primary_assignee"],
                    primary_assignee=fields["primary_assignee"],
                    category=fields["category"],
                    title=fields["title"],
                    content=fields["content"],
                    kanban_status=fields["kanban_status"],
                    priority=fields["priority"],
                    started_at=fields["started_at"],
                    closed_at=fields["closed_at"],
                    jira_issue_id=fields["jira_issue_id"],
                    jira_issue_key=fields["jira_issue_key"],
                    jira_url=fields["jira_url"],
                    jira_status=fields["jira_status"],
                    jira_status_category=fields.get("jira_status_category"),
                    jira_updated_at=fields["jira_updated_at"],
                    jira_epic=fields.get("jira_epic"),
                    jira_epic_key=fields.get("jira_epic_key"),
                    jira_epic_summary=fields.get("jira_epic_summary"),
                    jira_issue_type=fields.get("jira_issue_type"),
                    jira_parent_key=fields.get("jira_parent_key"),
                    jira_parent_summary=fields.get("jira_parent_summary"),
                    jira_components=fields.get("jira_components"),
                    jira_labels=fields.get("jira_labels"),
                    confluence_url=fields.get("confluence_url"),
                    jira_synced_at=now,
                    jira_watchers=[actor.username],
                    created_by=actor.username,
                )
                db.add(item)
                created += 1
        except Exception as exc:  # noqa: BLE001 - 한 이슈 실패가 전체를 막지 않도록
            logger.warning("Jira import 항목 실패 (%s): %s", issue.get("key"), exc)
            errors.append(f"{issue.get('key', '?')}: {str(exc)[:120]}")

    if not payload.dry_run:
        try:
            db.commit()
        except Exception as exc:  # noqa: BLE001
            db.rollback()
            return JiraImportResult(status="error", detail=f"저장 실패: {str(exc)[:200]}")
        audit_logger.record(
            db, action="work_item.jira_import", actor=actor,
            target_type="work_item", target_id=None,
            details={"scope": payload.scope, "created": created, "updated": updated, "skipped": skipped},
        )

    return JiraImportResult(
        status="ok",
        applied_jql=jql,
        imported=created,
        updated=updated,
        skipped=skipped,
        total=search.get("total", len(issues)),
        truncated=bool(search.get("truncated")),
        dry_run=payload.dry_run,
        errors=errors,
        items=preview[:200],
    )


# ── Jira 에서 추출한 Excel(.xlsx/.xls) / 복사·붙여넣기 가져오기 (미리보기 전용, 저장 없음) ──
_EXCEL_MAX_ROWS = 2000  # 과도한 업로드로부터의 안전장치
_EXCEL_HEADER_SCAN_ROWS = 5  # 헤더가 1행이 아닐 수 있어(제목행 등) 최대 이 행수까지 탐색


def _find_header_in_rows(rows: list[tuple]) -> tuple[Optional[tuple], int, dict[str, int], list[tuple]]:
    """rows 앞부분(최대 _EXCEL_HEADER_SCAN_ROWS 행)에서 key/summary 를 모두 가진 헤더 행을
    찾는다. 반환: (헤더 행 또는 None, 헤더의 rows 내 인덱스, 컬럼 인덱스 맵, 스캔한 후보 행들)."""
    scanned = rows[:_EXCEL_HEADER_SCAN_ROWS]
    for ridx, candidate in enumerate(scanned):
        normalized = [_norm_excel_header(h) for h in candidate]
        candidate_idx: dict[str, int] = {}
        for field, aliases in _EXCEL_HEADER_ALIASES.items():
            for i, h in enumerate(normalized):
                if h in aliases:
                    candidate_idx[field] = i
                    break
        if "key" in candidate_idx and "summary" in candidate_idx:
            return candidate, ridx, candidate_idx, scanned
    return None, -1, {}, scanned


def _extract_jira_rows(tables: list[list[tuple]], db: Session) -> JiraExcelImportResult:
    """표(시트/HTML 표/붙여넣은 TSV) 목록을 순서대로 살펴 key/summary 헤더를 가진 첫 표를
    찾아 JiraExcelRow 목록으로 변환한다. 파일 업로드(import_excel)와 복사·붙여넣기
    (import_paste) 가 공유하는 핵심 로직 — 표가 여러 개(예: Jira 요약 표 + 실제 이슈 표)
    있어도 순서대로 확인해 첫 번째로 헤더를 찾은 표를 사용한다."""
    tables = [t for t in tables if t]
    if not tables:
        return JiraExcelImportResult(status="error", detail="빈 파일입니다.")

    header_row: Optional[tuple] = None
    header_row_idx = -1
    col_idx: dict[str, int] = {}
    data_rows: list[tuple] = []
    scan_report: list[str] = []
    for t_idx, table_rows in enumerate(tables):
        h, hidx, cidx, scanned = _find_header_in_rows(table_rows)
        if h is not None:
            header_row, header_row_idx, col_idx = h, hidx, cidx
            data_rows = table_rows[hidx + 1:]
            break
        scanned_desc = "; ".join(
            f"{i + 1}행: {', '.join(str(v) for v in r if v) or '(빈 행)'}"
            for i, r in enumerate(scanned)
        )
        scan_report.append(f"[표 {t_idx + 1}] {scanned_desc}")

    if header_row is None:
        table_note = f"표 {len(tables)}개 확인 — " if len(tables) > 1 else ""
        return JiraExcelImportResult(
            status="error",
            detail=(
                f"필수 컬럼(key, summary)을 찾을 수 없습니다 (표마다 최대 "
                f"{_EXCEL_HEADER_SCAN_ROWS}행까지 확인, {table_note}각 표의 첫 행들: "
                + " / ".join(scan_report) + ")"
            ),
        )

    cfg = _get_config(db)
    base_url = (cfg.get("base_url") or "").rstrip("/")
    roster_by_lower = _build_assignee_roster(db)

    def cell(r: tuple, field: str) -> str:
        idx = col_idx.get(field)
        return _excel_cell_str(r[idx]) if idx is not None and idx < len(r) else ""

    out_rows: list[JiraExcelRow] = []
    matched = 0
    for r in data_rows:
        if r is None or all(v is None or v == "" for v in r):
            continue
        key = cell(r, "key")
        if not key:
            continue
        assignee_raw = cell(r, "assignee")
        assignee_name, is_matched = _match_excel_assignee(assignee_raw, roster_by_lower)
        if is_matched:
            matched += 1
        out_rows.append(JiraExcelRow(
            key=key,
            jira_url=f"{base_url}/browse/{key}" if base_url else None,
            summary=cell(r, "summary"),
            issue_type=cell(r, "issue_type"),
            status=cell(r, "status"),
            assignee_raw=assignee_raw,
            assignee_name=assignee_name,
            assignee_matched=is_matched,
            created=_excel_date_only(cell(r, "created")),
            resolved=cell(r, "resolved"),
            due_date=cell(r, "due_date"),
            environment=_strip_inline_html(cell(r, "environment")),
            description=_strip_inline_html(cell(r, "description")),
        ))
        if len(out_rows) >= _EXCEL_MAX_ROWS:
            break

    return JiraExcelImportResult(status="ok", total=len(out_rows), matched=matched, rows=out_rows)


@router.post("/import/excel", response_model=JiraExcelImportResult)
async def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Jira 에서 내려받은 이슈 목록 Excel(.xlsx/.xls) 을 파싱해 테이블로 보여준다.

    WorkItem 으로 저장하지 않는 순수 미리보기 기능 — 담당자 셀("이름 회사")에서
    이름을 추출해 등록된 담당자(Settings ▸ 담당자) 레지스트리와 매칭한다.
    """
    fname = (file.filename or "").lower()
    if not fname.endswith((".xlsx", ".xlsm", ".xls")):
        return JiraExcelImportResult(status="error", detail=".xlsx/.xls 파일만 업로드할 수 있습니다.")

    raw = await file.read()
    try:
        if fname.endswith(".xls"):
            if _looks_like_html(raw):
                # Jira '엑셀(전체 필드)' 내보내기 — 확장자만 .xls, 실제로는 HTML 표(여러 개일 수 있음).
                tables = _read_html_tables(raw)
            else:
                tables = [list(_read_xls_rows(raw))]
        else:
            wb = openpyxl.load_workbook(BytesIO(raw), data_only=True, read_only=True)
            ws = wb.active
            if ws is None:
                return JiraExcelImportResult(status="error", detail="시트를 찾을 수 없습니다.")
            tables = [list(ws.iter_rows(values_only=True))]
    except Exception as exc:  # noqa: BLE001
        return JiraExcelImportResult(status="error", detail=f"엑셀 파일을 읽을 수 없습니다: {str(exc)[:150]}")

    return _extract_jira_rows(tables, db)


@router.post("/import/paste", response_model=JiraExcelImportResult)
async def import_paste(
    payload: JiraExcelPasteRequest,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """엑셀/Jira 표를 복사해 그대로 붙여넣은 텍스트를 파싱한다(파일 업로드 없이).
    브라우저에서 표를 드래그로 복사하면 클립보드에 탭(TSV) 구분 텍스트가 담기는 것을
    이용 — 파일 업로드(import_excel)와 동일한 헤더 탐색/매칭 로직을 공유한다."""
    text = payload.text.replace("\r\n", "\n").replace("\r", "\n")
    lines = [ln for ln in text.split("\n") if ln.strip() != ""]
    if not lines:
        return JiraExcelImportResult(status="error", detail="붙여넣은 내용이 비어 있습니다.")
    rows: list[tuple] = [tuple(cell.strip() for cell in ln.split("\t")) for ln in lines]
    return _extract_jira_rows([rows], db)


@router.post("/import/excel/save", response_model=JiraImportResult)
async def import_excel_save(
    payload: JiraExcelSaveRequest,
    db: Session = Depends(get_db),
    actor: User = Depends(require_operator),
):
    """import_excel/import_paste 로 미리 확인한 행을 실제 업무 관리 게시판(work_items)에
    저장한다. 파일을 다시 읽지 않고 프론트가 들고 있던 미리보기 rows 를 그대로 받는다.

    라이브 JQL 가져오기(POST /import)는 `jira_issue_id`(Jira 내부 불변 ID)로 dedup 하지만,
    엑셀에는 그 값이 없으므로 여기서는 `jira_issue_key`(예: PROJ-123)로 dedup 한다 — 같은
    이슈를 라이브로 먼저 가져왔다면 jira_issue_key 가 이미 채워져 있어 그 레코드를 그대로
    갱신한다(반대로 엑셀로 먼저 저장한 뒤 라이브로 가져오면 jira_issue_id 매칭에는 안 걸려
    새 레코드가 생길 수 있음 — 라이브 가져오기 dedup 정책은 이번 변경 범위 밖).
    """
    created = updated = skipped = 0
    errors: list[str] = []
    preview: list[JiraImportItemPreview] = []
    now = datetime.utcnow()

    for row in payload.rows:
        key = (row.key or "").strip()
        if not key:
            skipped += 1
            continue
        try:
            wtype, type_label = map_issue_type(row.issue_type)
            kanban = _map_excel_status_to_kanban(row.status)
            summary = (row.summary or "").strip()
            title = f"{key} {summary}".strip()[:200]
            content = row.description if (row.description or "").strip() else (summary or key)
            started_at = _parse_excel_date(row.created) or now
            closed_at = _parse_excel_date(row.resolved) if kanban == "done" else None
            # assignee_name 은 미리보기 단계(_match_excel_assignee)에서 이미 담당자 레지스트리와
            # 매칭 시도된 값 — 매칭 실패해도 원본 첫 토큰이 담겨 있어 빈 문자열이 아닌 한 그대로 쓴다.
            assignee_name = (row.assignee_name or "").strip() or "(미할당)"

            existing = db.query(WorkItem).filter(WorkItem.jira_issue_key == key).first()
            action = "update" if existing else "create"
            preview.append(JiraImportItemPreview(
                jira_key=key, title=title, kanban_status=kanban, action=action,
            ))

            if existing:
                existing.title = title
                existing.content = content
                existing.kanban_status = kanban
                existing.jira_status = row.status
                existing.jira_url = row.jira_url
                existing.jira_synced_at = now
                if closed_at and not existing.closed_at:
                    existing.closed_at = closed_at
                if not (existing.primary_assignee or "").strip() or existing.primary_assignee == "(미할당)":
                    existing.primary_assignee = assignee_name
                    existing.assignee = assignee_name
                watchers = list(existing.jira_watchers or [])
                if actor.username not in watchers:
                    watchers.append(actor.username)
                existing.jira_watchers = watchers
                updated += 1
            else:
                item = WorkItem(
                    type=wtype,
                    type_label=type_label,
                    assignee=assignee_name,
                    primary_assignee=assignee_name,
                    category="Jira",
                    title=title,
                    content=content,
                    kanban_status=kanban,
                    started_at=started_at,
                    closed_at=closed_at,
                    jira_issue_key=key,
                    jira_url=row.jira_url,
                    jira_status=row.status,
                    jira_synced_at=now,
                    jira_watchers=[actor.username],
                    created_by=actor.username,
                )
                db.add(item)
                created += 1
        except Exception as exc:  # noqa: BLE001 - 한 행 실패가 전체 저장을 막지 않도록
            logger.warning("Jira Excel 저장 항목 실패 (%s): %s", key, exc)
            errors.append(f"{key}: {str(exc)[:120]}")

    try:
        db.commit()
    except Exception as exc:  # noqa: BLE001
        db.rollback()
        return JiraImportResult(status="error", detail=f"저장 실패: {str(exc)[:200]}")

    audit_logger.record(
        db, action="work_item.jira_excel_import", actor=actor,
        target_type="work_item", target_id=None,
        details={"created": created, "updated": updated, "skipped": skipped},
    )

    return JiraImportResult(
        status="ok",
        imported=created,
        updated=updated,
        skipped=skipped,
        total=len(payload.rows),
        errors=errors,
        items=preview[:50],
    )


# ── 양방향 push: PEP 상태 → Jira 반영 (Phase 2) ─────────────────────────────────
@router.post("/push/{item_id}", response_model=JiraPushResult)
async def push_to_jira(
    item_id: str,
    payload: JiraPushRequest,
    db: Session = Depends(get_db),
    actor: User = Depends(require_operator),
):
    item = db.query(WorkItem).filter(WorkItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="업무를 찾을 수 없습니다.")
    if not item.jira_issue_key:
        return JiraPushResult(status="not_linked", detail="Jira 와 연결되지 않은 업무입니다.")

    cfg = _get_config(db)
    if not cfg.get("base_url") or not cfg.get("enabled", False):
        return JiraPushResult(status="error", detail="Jira 연동이 비활성화되었거나 URL 미설정.")
    token, _auth_type = _user_credential(db, actor.username)
    if not token:
        return JiraPushResult(status="error", detail="내 Jira 인증(PAT 또는 세션 쿠키)이 등록되지 않았습니다 (설정 > 연동).")

    key = item.jira_issue_key
    # 세션 만료(401) 시 저장된 SSO 로그인 정보로 자동 재로그인 포함.
    svc, _myself = await _jira_service_verified(db, actor, cfg)
    if svc is None:
        return JiraPushResult(status="error", detail="내 Jira 인증(PAT 또는 세션 쿠키)이 등록되지 않았습니다 (설정 > 연동).")

    # 1) 현재 Jira 상태 + updated 조회 (충돌 감지)
    got = await svc.get_issue(key, fields=["status", "updated"])
    if got.get("status") != "ok":
        return JiraPushResult(status=got.get("status", "error"), detail=got.get("detail", "이슈 조회 실패"))
    jfields = (got["issue"].get("fields") or {})
    jira_updated = parse_jira_dt(jfields.get("updated"))
    if (not payload.force) and item.jira_updated_at and jira_updated and jira_updated > item.jira_updated_at:
        return JiraPushResult(
            status="conflict",
            detail="Jira 쪽이 더 최근에 변경되었습니다. 덮어쓰려면 강제 반영을 선택하세요.",
            jira_status=(jfields.get("status") or {}).get("name"),
        )

    # 1.5) 값 필드 편집 반영 (제목/설명/우선순위). assignee 는 PEP 담당자명 ↔ Jira username
    #      역매핑이 불안정해 제외한다. summary/description 은 저위험이라 한 번에 PUT 하고,
    #      priority 는 프로젝트별 우선순위 스킴 차이로 실패할 수 있어 별도 best-effort PUT.
    fields_updated: list[str] = []
    field_errors: list[str] = []
    if payload.push_fields:
        core: dict = {}
        summary = strip_issue_key_prefix(item.title, key)
        if summary:
            core["summary"] = summary[:255]
        if item.content is not None:
            core["description"] = item.content or ""
        if core:
            upd = await svc.update_issue(key, core)
            if upd.get("status") == "ok":
                fields_updated.extend(core.keys())
            else:
                field_errors.append(upd.get("detail", "제목/설명 반영 실패"))
        jira_priority = PEP_PRIORITY_TO_JIRA.get((item.priority or "medium").lower())
        if jira_priority:
            pres = await svc.update_issue(key, {"priority": {"name": jira_priority}})
            if pres.get("status") == "ok":
                fields_updated.append("priority")
            else:
                field_errors.append(f"priority({jira_priority}): {pres.get('detail', '반영 실패')}")

    # 2) 상태 transition
    transitioned = False
    desired_cat = KANBAN_TO_CATEGORY.get(item.kanban_status or "todo", "new")
    current_cat = ((jfields.get("status") or {}).get("statusCategory") or {}).get("key", "")
    if current_cat != desired_cat:
        tr = await svc.get_transitions(key)
        if tr.get("status") != "ok":
            return JiraPushResult(status=tr.get("status", "error"), detail=tr.get("detail", "transition 조회 실패"))
        match = next((t for t in tr["transitions"] if t.get("to_category") == desired_cat), None)
        if not match:
            names = [t.get("name", "") for t in tr["transitions"]]
            return JiraPushResult(
                status="error",
                detail=f"'{item.kanban_status}' 에 해당하는 Jira transition 이 없습니다. 가용: {', '.join(names) or '없음'}",
                available_transitions=names,
            )
        res = await svc.do_transition(key, match["id"])
        if res.get("status") != "ok":
            return JiraPushResult(status=res.get("status", "error"), detail=res.get("detail", "transition 실패"))
        transitioned = True

    # 3) 코멘트 (선택)
    comment_added = False
    if payload.comment and payload.comment.strip():
        cres = await svc.add_comment(key, payload.comment.strip())
        comment_added = cres.get("status") == "ok"

    # 4) 로컬 동기화 메타 갱신 (재조회)
    after = await svc.get_issue(key, fields=["status", "updated"])
    now = datetime.utcnow()
    new_status_name = None
    if after.get("status") == "ok":
        af = after["issue"].get("fields") or {}
        new_status_name = (af.get("status") or {}).get("name")
        item.jira_status = new_status_name
        item.jira_updated_at = parse_jira_dt(af.get("updated")) or item.jira_updated_at
    item.jira_synced_at = now
    db.commit()

    audit_logger.record(
        db, action="work_item.jira_push", actor=actor,
        target_type="work_item", target_id=str(item.id),
        details={
            "key": key, "transitioned": transitioned, "comment": comment_added,
            "fields": fields_updated, "force": payload.force,
        },
    )

    changed = transitioned or comment_added or bool(fields_updated)
    if changed:
        parts = []
        if fields_updated:
            _labels = {"summary": "제목", "description": "설명", "priority": "우선순위"}
            parts.append("필드(" + ", ".join(_labels.get(f, f) for f in fields_updated) + ")")
        if transitioned:
            parts.append("상태")
        if comment_added:
            parts.append("코멘트")
        detail = "Jira 반영 완료: " + " · ".join(parts)
    else:
        detail = "이미 동기화 상태입니다."
    if field_errors:
        detail += " (일부 실패: " + "; ".join(field_errors[:3]) + ")"
    return JiraPushResult(
        status="ok", detail=detail, transitioned=transitioned,
        comment_added=comment_added, fields_updated=fields_updated, field_errors=field_errors,
        jira_status=new_status_name or (jfields.get("status") or {}).get("name"),
    )
